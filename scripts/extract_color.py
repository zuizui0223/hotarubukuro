#!/usr/bin/env python3
"""Deterministic petal-colour extraction from images or an XLSX workbook.

Version 2 replaces the previous one-cluster K-means calculation (which was
only an arithmetic mean) with an explicit, inspectable mask and mean, median,
and joint CIELAB modal-peak summaries.  For RGBA images, alpha is the primary
foreground bound; HSV petal candidates are always restricted to visible
pixels.  The primary R/G/B statistic defaults to the channel-wise median; the
modal estimator is additive until it has been empirically validated.

Excel "Place in Cell" images are not exposed by openpyxl.  They are resolved
directly from OOXML using the following chain::

    worksheet cell vm -> metadata.xml -> rich value -> richValueRel -> media

Ordinary DrawingML images anchored to cells are supported as well.  Workbook
rows are joined to images by their cell and then emitted with an immutable
``observation_id``; positional joins to a separate table are never used.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import math
import os
import posixpath
import re
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, MutableMapping, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

import numpy as np
from PIL import Image, ImageDraw, ImageFilter, ImageOps


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
RICH_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
RICH_REL_NS = "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
EXTRACTION_VERSION = "2.2.2"
PEAK_LAB_BIN_WIDTH = 5.0
PEAK_DELTA_E_RADIUS = 12.0
PEAK_SECOND_MIN_DELTA_E = 24.0
PEAK_MULTIMODAL_THRESHOLD = 0.5
PEAK_LOW_FRACTION_THRESHOLD = 0.2
EXPOSURE_FILTERED_MIN_PIXELS = 50
EXPOSURE_FILTERED_MIN_FRACTION = 0.25
MASK_LARGE_COMPONENT_MIN_FRACTION = 0.1
NEAR_WHITE_FRACTION_THRESHOLD = 0.1
ALL_CHANNELS_CLIPPED_THRESHOLD = 0.01
DARK_SHADOW_THRESHOLD = 0.25
EXCLUDED_WARM_THRESHOLD = 0.05
COOL_BLUE_THRESHOLD = 0.05
ALPHA_HSV_PEAK_DISAGREEMENT_THRESHOLD = 10.0
LOW_MASK_COVERAGE_THRESHOLD = 0.8

OUTPUT_COLUMNS = [
    "observation_id",
    "photo_id",
    "source_row",
    "source_sheet",
    "source_image",
    "source_path",
    "url",
    "date",
    "latitude",
    "longitude",
    "legacy_R",
    "legacy_G",
    "legacy_B",
    "delta_R_vs_legacy",
    "delta_G_vs_legacy",
    "delta_B_vs_legacy",
    "delta_rgb_euclidean_vs_legacy",
    "image_sha256",
    "duplicate_image_sha256",
    "image_width",
    "image_height",
    "has_alpha",
    "alpha_threshold",
    "source_exif_present",
    "camera_exif_present",
    "white_balance_metadata_present",
    "source_srgb_chunk_present",
    "source_icc_profile_present",
    "source_colour_space",
    "neutral_reference_available",
    "illumination_correction_status",
    "visible_pixels",
    "mask_pixels",
    "mask_fraction",
    "mask_fraction_visible",
    "low_mask_coverage_threshold",
    "mask_source",
    "alpha_segmentation_source",
    "morph_kernel_px",
    "morph_kernel_short_side_fraction",
    "mask_component_count",
    "mask_large_component_count",
    "mask_largest_component_fraction",
    "possible_multiple_flowers",
    "alpha_component_count",
    "alpha_large_component_count",
    "alpha_largest_component_fraction",
    "possible_multiple_components",
    "excluded_warm_fraction_visible",
    "cool_blue_fraction_visible",
    "warm_sensitivity_mask_fraction_visible",
    "near_white_fraction",
    "all_channels_clipped_fraction",
    "any_channel_clipped_fraction",
    "alpha_all_channels_clipped_fraction",
    "alpha_any_channel_clipped_fraction",
    "possible_overexposure",
    "dark_shadow_fraction",
    "alpha_dark_shadow_fraction",
    "mean_R",
    "mean_G",
    "mean_B",
    "median_R",
    "median_G",
    "median_B",
    "hsv_peak_R",
    "hsv_peak_G",
    "hsv_peak_B",
    "hsv_peak_L",
    "hsv_peak_a",
    "hsv_peak_b",
    "hsv_peak_pixels",
    "hsv_peak_fraction",
    "hsv_peak_dispersion_deltaE",
    "hsv_second_peak_ratio",
    "hsv_multimodal_colour",
    "hsv_exposure_filtered_peak_R",
    "hsv_exposure_filtered_peak_G",
    "hsv_exposure_filtered_peak_B",
    "hsv_exposure_filtered_peak_L",
    "hsv_exposure_filtered_peak_a",
    "hsv_exposure_filtered_peak_b",
    "hsv_exposure_filtered_peak_pixels",
    "hsv_exposure_filtered_peak_fraction",
    "hsv_exposure_filtered_peak_dispersion_deltaE",
    "hsv_exposure_filtered_second_peak_ratio",
    "hsv_exposure_filtered_multimodal_colour",
    "exposure_filtered_pixels",
    "exposure_filtered_fraction",
    "exposure_filtered_min_pixels",
    "exposure_filtered_min_fraction",
    "exposure_filtered_peak_available",
    "alpha_peak_R",
    "alpha_peak_G",
    "alpha_peak_B",
    "alpha_peak_L",
    "alpha_peak_a",
    "alpha_peak_b",
    "alpha_peak_pixels",
    "alpha_peak_fraction",
    "alpha_peak_dispersion_deltaE",
    "alpha_second_peak_ratio",
    "alpha_multimodal_colour",
    "alpha_hsv_peak_deltaE",
    "alpha_hsv_peak_disagreement_threshold",
    "peak_method",
    "peak_lab_bin_width",
    "peak_delta_e_radius",
    "peak_second_min_delta_e",
    "peak_multimodal_threshold",
    "peak_low_fraction_threshold",
    "R",
    "G",
    "B",
    "white_balance_applied",
    "white_balance_method",
    "white_balance_reliability",
    "white_balance_note",
    "color_statistic",
    "primary_colour_method",
    "candidate_colour_methods",
    "extraction_version",
    "processed_at",
    "qc_status",
    "qc_flags",
    "qc_note",
    "qc_sampled",
    "mask_path",
    "overlay_path",
    "qc_panel_path",
]


@dataclass(frozen=True)
class SourceRecord:
    """One photograph observation and its provenance."""

    observation_id: str
    photo_id: str
    source_row: Optional[int] = None
    source_sheet: Optional[str] = None
    source_image: Optional[str] = None
    url: Any = None
    date: Any = None
    latitude: Any = None
    longitude: Any = None
    legacy_R: Any = None
    legacy_G: Any = None
    legacy_B: Any = None
    file_path: Optional[Path] = None
    workbook_path: Optional[Path] = None
    zip_member: Optional[str] = None


@dataclass
class Extraction:
    """Decoded image, mask, and scalar colour metrics."""

    rgb: np.ndarray
    alpha: np.ndarray
    mask: np.ndarray
    metrics: Dict[str, Any]


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _read_xml(archive: zipfile.ZipFile, member: str) -> ET.Element:
    try:
        return ET.fromstring(archive.read(member))
    except KeyError as exc:
        raise ValueError("Workbook is missing required OOXML part: %s" % member) from exc
    except ET.ParseError as exc:
        raise ValueError("Invalid XML in workbook part: %s" % member) from exc


def _relationship_part(source_part: str) -> str:
    return posixpath.join(
        posixpath.dirname(source_part),
        "_rels",
        posixpath.basename(source_part) + ".rels",
    )


def _resolve_target(source_part: str, target: str) -> str:
    resolved = posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))
    if resolved.startswith("../") or resolved == "..":
        raise ValueError("OOXML relationship escapes the workbook archive: %s" % target)
    return resolved.lstrip("/")


def _relationships(archive: zipfile.ZipFile, source_part: str) -> Dict[str, str]:
    rel_part = _relationship_part(source_part)
    if rel_part not in archive.namelist():
        return {}
    root = _read_xml(archive, rel_part)
    out: Dict[str, str] = {}
    for rel in root.findall("{%s}Relationship" % PKG_REL_NS):
        rel_id = rel.get("Id")
        target = rel.get("Target")
        if rel_id and target and rel.get("TargetMode") != "External":
            out[rel_id] = _resolve_target(source_part, target)
    return out


def _column_letter(column: int) -> str:
    """Return an Excel column label for a one-based column number."""

    if column < 1:
        raise ValueError("Excel column number must be positive")
    letters: List[str] = []
    while column:
        column, remainder = divmod(column - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def _sheet_parts(archive: zipfile.ZipFile) -> List[Tuple[str, str]]:
    workbook_part = "xl/workbook.xml"
    root = _read_xml(archive, workbook_part)
    rels = _relationships(archive, workbook_part)
    sheets: List[Tuple[str, str]] = []
    for sheet in root.findall("{%s}sheets/{%s}sheet" % (MAIN_NS, MAIN_NS)):
        name = sheet.get("name")
        rel_id = sheet.get("{%s}id" % REL_NS)
        if name and rel_id in rels:
            sheets.append((name, rels[rel_id]))
    if not sheets:
        raise ValueError("Workbook contains no readable worksheets")
    return sheets


def _selected_sheet_part(
    archive: zipfile.ZipFile, sheet_name: Optional[str]
) -> Tuple[str, str]:
    sheets = _sheet_parts(archive)
    if sheet_name is None:
        return sheets[0]
    for name, part in sheets:
        if name == sheet_name:
            return name, part
    raise ValueError(
        "Worksheet %r not found; available sheets: %s"
        % (sheet_name, ", ".join(name for name, _ in sheets))
    )


def _rich_image_map(
    archive: zipfile.ZipFile, sheet_part: str
) -> Dict[str, str]:
    """Return ``cell reference -> media member`` for rich in-cell images."""

    sheet_root = _read_xml(archive, sheet_part)
    cell_vm: Dict[str, int] = {}
    for cell in sheet_root.findall(".//{%s}c" % MAIN_NS):
        ref = cell.get("r")
        vm = cell.get("vm")
        if ref and vm:
            try:
                cell_vm[ref] = int(vm)
            except ValueError as exc:
                raise ValueError("Invalid rich-value metadata index at cell %s" % ref) from exc
    if not cell_vm:
        return {}

    required = {
        "xl/metadata.xml",
        "xl/richData/rdrichvalue.xml",
        "xl/richData/rdrichvaluestructure.xml",
        "xl/richData/richValueRel.xml",
        "xl/richData/_rels/richValueRel.xml.rels",
    }
    missing = sorted(required.difference(archive.namelist()))
    if missing:
        raise ValueError(
            "Worksheet uses rich values but workbook lacks: %s" % ", ".join(missing)
        )

    metadata = _read_xml(archive, "xl/metadata.xml")
    value_metadata: List[int] = []
    value_root = metadata.find("{%s}valueMetadata" % MAIN_NS)
    if value_root is not None:
        for block in value_root.findall("{%s}bk" % MAIN_NS):
            rc = block.find("{%s}rc" % MAIN_NS)
            if rc is None or rc.get("v") is None:
                value_metadata.append(-1)
            else:
                value_metadata.append(int(rc.get("v", "-1")))

    rich_indexes: List[int] = []
    future = None
    for candidate in metadata.findall("{%s}futureMetadata" % MAIN_NS):
        if candidate.get("name") == "XLRICHVALUE":
            future = candidate
            break
    if future is not None:
        for block in future.findall("{%s}bk" % MAIN_NS):
            marker = block.find(".//{%s}rvb" % RICH_NS)
            rich_indexes.append(int(marker.get("i", "-1")) if marker is not None else -1)

    structures_root = _read_xml(archive, "xl/richData/rdrichvaluestructure.xml")
    structures: List[List[str]] = []
    for structure in structures_root.findall("{%s}s" % RICH_NS):
        structures.append(
            [key.get("n", "") for key in structure.findall("{%s}k" % RICH_NS)]
        )

    values_root = _read_xml(archive, "xl/richData/rdrichvalue.xml")
    rich_values = values_root.findall("{%s}rv" % RICH_NS)

    rel_list_root = _read_xml(archive, "xl/richData/richValueRel.xml")
    relation_ids = [
        rel.get("{%s}id" % REL_NS)
        for rel in rel_list_root.findall("{%s}rel" % RICH_REL_NS)
    ]
    relation_targets = _relationships(archive, "xl/richData/richValueRel.xml")

    out: Dict[str, str] = {}
    for cell_ref, vm in cell_vm.items():
        vm_index = vm - 1  # worksheet vm is one-based
        if not (0 <= vm_index < len(value_metadata)):
            raise ValueError("Cell %s references missing valueMetadata %s" % (cell_ref, vm))
        future_index = value_metadata[vm_index]
        if not (0 <= future_index < len(rich_indexes)):
            raise ValueError("Cell %s references invalid rich metadata" % cell_ref)
        rich_index = rich_indexes[future_index]
        if not (0 <= rich_index < len(rich_values)):
            raise ValueError("Cell %s references missing rich value" % cell_ref)

        rich_value = rich_values[rich_index]
        structure_index = int(rich_value.get("s", "0"))
        value_nodes = rich_value.findall("{%s}v" % RICH_NS)
        key_names = structures[structure_index] if 0 <= structure_index < len(structures) else []
        relation_position = 0
        for index, key_name in enumerate(key_names):
            if key_name.endswith("LocalImageIdentifier"):
                relation_position = index
                break
        if relation_position >= len(value_nodes) or value_nodes[relation_position].text is None:
            raise ValueError("Rich image at cell %s has no relationship index" % cell_ref)
        relation_index = int(value_nodes[relation_position].text)
        if not (0 <= relation_index < len(relation_ids)):
            raise ValueError("Rich image at cell %s has invalid relationship index" % cell_ref)
        relation_id = relation_ids[relation_index]
        if not relation_id or relation_id not in relation_targets:
            raise ValueError("Rich image relationship missing for cell %s" % cell_ref)
        target = relation_targets[relation_id]
        if target not in archive.namelist():
            raise ValueError("Rich image media is missing for cell %s: %s" % (cell_ref, target))
        out[cell_ref] = target
    return out


def _drawing_image_map(
    archive: zipfile.ZipFile, sheet_part: str
) -> Dict[str, str]:
    """Return ``cell reference -> media member`` for anchored DrawingML images."""

    sheet_root = _read_xml(archive, sheet_part)
    sheet_rels = _relationships(archive, sheet_part)
    drawing_nodes = sheet_root.findall("{%s}drawing" % MAIN_NS)
    if not drawing_nodes:
        return {}

    out: Dict[str, str] = {}
    for drawing_node in drawing_nodes:
        rel_id = drawing_node.get("{%s}id" % REL_NS)
        if not rel_id or rel_id not in sheet_rels:
            continue
        drawing_part = sheet_rels[rel_id]
        drawing_root = _read_xml(archive, drawing_part)
        drawing_rels = _relationships(archive, drawing_part)
        for anchor in list(drawing_root):
            if anchor.tag not in {
                "{%s}oneCellAnchor" % XDR_NS,
                "{%s}twoCellAnchor" % XDR_NS,
            }:
                continue
            start = anchor.find("{%s}from" % XDR_NS)
            if start is None:
                continue
            row_node = start.find("{%s}row" % XDR_NS)
            col_node = start.find("{%s}col" % XDR_NS)
            blip = anchor.find(".//{%s}blip" % DRAWING_NS)
            if row_node is None or col_node is None or blip is None:
                continue
            rel_id = blip.get("{%s}embed" % REL_NS)
            if not rel_id or rel_id not in drawing_rels:
                continue
            cell_ref = "%s%d" % (
                _column_letter(int(col_node.text or "0") + 1),
                int(row_node.text or "0") + 1,
            )
            target = drawing_rels[rel_id]
            if target not in archive.namelist():
                raise ValueError("Drawing image media is missing: %s" % target)
            if cell_ref in out and out[cell_ref] != target:
                raise ValueError("Multiple images are anchored to cell %s" % cell_ref)
            out[cell_ref] = target
    return out


def workbook_image_map(
    workbook_path: Path, sheet_name: Optional[str] = None
) -> Tuple[str, Dict[str, str]]:
    """Map in-cell and anchored workbook images to worksheet cells."""

    with zipfile.ZipFile(workbook_path) as archive:
        selected_name, sheet_part = _selected_sheet_part(archive, sheet_name)
        rich = _rich_image_map(archive, sheet_part)
        drawings = _drawing_image_map(archive, sheet_part)
        for cell_ref, target in drawings.items():
            if cell_ref in rich and rich[cell_ref] != target:
                raise ValueError("Cell %s has conflicting rich and drawing images" % cell_ref)
            rich[cell_ref] = target
        return selected_name, rich


def _normalise_header(value: Any) -> str:
    return str(value).strip().casefold() if value is not None else ""


def _normalise_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped and stripped != "#VALUE!" else None
    return value


def processing_timestamp() -> str:
    """Return one UTC run timestamp, honouring reproducible-build convention."""

    epoch = os.environ.get("SOURCE_DATE_EPOCH")
    if epoch is not None:
        try:
            moment = datetime.fromtimestamp(int(epoch), tz=timezone.utc)
        except (OverflowError, ValueError) as exc:
            raise ValueError("SOURCE_DATE_EPOCH must be an integer Unix timestamp") from exc
    else:
        moment = datetime.now(timezone.utc)
    return moment.replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _first_value(values: Mapping[str, Any], aliases: Sequence[str]) -> Any:
    for alias in aliases:
        key = alias.casefold()
        if key in values:
            return _normalise_scalar(values[key])
    return None


def _resolve_column(
    headers: Mapping[str, int], requested: str, *, required: bool
) -> Optional[int]:
    key = requested.strip().casefold()
    if key in headers:
        return headers[key]
    if re.fullmatch(r"[A-Za-z]{1,3}", requested.strip()):
        try:
            from openpyxl.utils import column_index_from_string

            return column_index_from_string(requested.strip())
        except (ImportError, ValueError):
            pass
    if required:
        raise ValueError(
            "Column %r not found; available headers: %s"
            % (requested, ", ".join(sorted(headers)))
        )
    return None


def collect_workbook_records(
    workbook_path: Path,
    *,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    image_column: str = "petal",
    id_column: str = "observation_id",
) -> List[SourceRecord]:
    """Read workbook metadata and associate each data row with its cell image."""

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Workbook input requires openpyxl; install .[excel]") from exc

    selected_name, image_map = workbook_image_map(workbook_path, sheet_name)
    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=False, keep_links=True
    )
    try:
        worksheet = workbook[selected_name]
        header_cells = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                min_col=1,
                max_col=worksheet.max_column,
            )
        )
        headers: Dict[str, int] = {}
        display_headers: Dict[int, str] = {}
        for column, cell in enumerate(header_cells, start=1):
            normalised = _normalise_header(cell.value)
            if not normalised:
                continue
            if normalised in headers:
                raise ValueError("Duplicate workbook header: %s" % cell.value)
            headers[normalised] = column
            display_headers[column] = normalised

        image_col = _resolve_column(headers, image_column, required=True)
        id_col = _resolve_column(headers, id_column, required=False)
        assert image_col is not None

        records: List[SourceRecord] = []
        for excel_row, row in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ),
            start=header_row + 1,
        ):
            row_values: Dict[str, Any] = {}
            for column, cell in enumerate(row, start=1):
                header = display_headers.get(column)
                if header:
                    row_values[header] = cell.value
            cell_ref = "%s%d" % (get_column_letter(image_col), excel_row)
            member = image_map.get(cell_ref)
            if not member and not any(
                _normalise_scalar(value) is not None for value in row_values.values()
            ):
                continue

            raw_id = row[id_col - 1].value if id_col is not None else None
            observation_id = str(_normalise_scalar(raw_id) or "obs-%06d" % (excel_row - header_row))
            existing_photo_id = _first_value(row_values, ("photo_id", "image_id"))
            photo_id = str(
                existing_photo_id
                or (Path(member).stem if member else "workbook-row-%d" % excel_row)
            )
            source_image = (
                "%s#%s" % (workbook_path.name, member) if member else None
            )
            records.append(
                SourceRecord(
                    observation_id=observation_id,
                    photo_id=photo_id,
                    source_row=excel_row,
                    source_sheet=selected_name,
                    source_image=source_image,
                    url=_first_value(row_values, ("url", "source_url")),
                    date=_first_value(row_values, ("date", "collection_date")),
                    latitude=_first_value(row_values, ("latitude", "lat")),
                    longitude=_first_value(row_values, ("longitude", "lon", "long")),
                    legacy_R=_first_value(row_values, ("r", "legacy_r")),
                    legacy_G=_first_value(row_values, ("g", "legacy_g")),
                    legacy_B=_first_value(row_values, ("b", "legacy_b")),
                    workbook_path=workbook_path,
                    zip_member=member,
                )
            )
    finally:
        workbook.close()
    _validate_observation_ids(records)
    return records


def collect_directory_records(input_dir: Path) -> List[SourceRecord]:
    """Collect supported image files in stable relative-path order."""

    if not input_dir.is_dir():
        raise ValueError("Input directory does not exist: %s" % input_dir)
    paths = [
        path
        for path in input_dir.rglob("*")
        if path.is_file() and path.suffix.casefold() in IMAGE_SUFFIXES
    ]
    paths.sort(
        key=lambda path: (
            path.relative_to(input_dir).as_posix().casefold(),
            path.relative_to(input_dir).as_posix(),
        )
    )
    records: List[SourceRecord] = []
    for path in paths:
        relative = path.relative_to(input_dir)
        observation_id = relative.with_suffix("").as_posix()
        records.append(
            SourceRecord(
                observation_id=observation_id,
                photo_id=observation_id,
                source_image=relative.as_posix(),
                file_path=path,
            )
        )
    _validate_observation_ids(records)
    return records


def _validate_observation_ids(records: Sequence[SourceRecord]) -> None:
    seen: Dict[str, int] = {}
    duplicates: List[str] = []
    for record in records:
        if not record.observation_id.strip():
            raise ValueError("observation_id cannot be empty")
        if record.observation_id in seen:
            duplicates.append(record.observation_id)
        seen[record.observation_id] = seen.get(record.observation_id, 0) + 1
    if duplicates:
        raise ValueError(
            "Duplicate observation_id values: %s" % ", ".join(sorted(set(duplicates)))
        )


def _decode_image(
    image_bytes: bytes,
) -> Tuple[np.ndarray, np.ndarray, bool, bool, bool, bool, bool, bool]:
    try:
        with Image.open(io.BytesIO(image_bytes)) as opened:
            exif = opened.getexif()
            source_exif_present = bool(exif)
            source_icc_profile_present = bool(opened.info.get("icc_profile"))
            exif_values: Dict[int, Any] = dict(exif)
            for ifd_tag in (34665, 34853):
                try:
                    exif_values.update(dict(exif.get_ifd(ifd_tag)))
                except (AttributeError, KeyError, TypeError, ValueError):
                    pass
            camera_tags = {
                271,  # Make
                272,  # Model
                33434,  # ExposureTime
                33437,  # FNumber
                34850,  # ExposureProgram
                34855,  # ISOSpeedRatings
                36867,  # DateTimeOriginal
                37377, 37378, 37380, 37383, 37385, 37386,
                41987,  # WhiteBalance
            }
            white_balance_tags = {41987, 50728, 50729}
            camera_exif_present = any(tag in exif_values for tag in camera_tags)
            white_balance_metadata_present = any(
                tag in exif_values for tag in white_balance_tags
            )
            source_srgb_chunk_present = bool(
                "srgb" in opened.info
                or "sRGB" in opened.info
                or exif_values.get(40961) == 1
            )
            transposed = ImageOps.exif_transpose(opened)
            original_has_alpha = transposed.mode in {"RGBA", "LA", "PA"} or (
                "transparency" in transposed.info
            )
            rgba = np.asarray(transposed.convert("RGBA"), dtype=np.uint8).copy()
    except Exception as exc:
        raise ValueError("Unreadable image: %s" % exc) from exc
    if rgba.ndim != 3 or rgba.shape[2] != 4 or rgba.size == 0:
        raise ValueError("Decoded image has an invalid shape")
    return (
        rgba[:, :, :3],
        rgba[:, :, 3],
        original_has_alpha,
        source_exif_present,
        camera_exif_present,
        white_balance_metadata_present,
        source_srgb_chunk_present,
        source_icc_profile_present,
    )


def _rgb_to_opencv_hsv(rgb: np.ndarray) -> np.ndarray:
    """Convert uint8 RGB to OpenCV-scaled HSV without requiring OpenCV.

    Hue uses 0..179 and saturation/value use 0..255, matching the thresholds
    traditionally used by this project.  Achromatic pixels receive hue zero.
    """

    values = rgb.astype(np.float64) / 255.0
    red, green, blue = values[:, :, 0], values[:, :, 1], values[:, :, 2]
    maximum = np.max(values, axis=2)
    minimum = np.min(values, axis=2)
    delta = maximum - minimum

    hue_degrees = np.zeros_like(maximum)
    chromatic = delta > 0
    red_max = chromatic & (maximum == red)
    green_max = chromatic & (maximum == green) & ~red_max
    blue_max = chromatic & ~(red_max | green_max)
    hue_degrees[red_max] = (
        60.0 * ((green[red_max] - blue[red_max]) / delta[red_max])
    ) % 360.0
    hue_degrees[green_max] = 60.0 * (
        (blue[green_max] - red[green_max]) / delta[green_max] + 2.0
    )
    hue_degrees[blue_max] = 60.0 * (
        (red[blue_max] - green[blue_max]) / delta[blue_max] + 4.0
    )

    saturation = np.zeros_like(maximum)
    nonblack = maximum > 0
    saturation[nonblack] = delta[nonblack] / maximum[nonblack]
    hsv = np.empty_like(rgb)
    hsv[:, :, 0] = np.clip(np.rint(hue_degrees / 2.0), 0, 179).astype(np.uint8)
    hsv[:, :, 1] = np.clip(np.rint(saturation * 255.0), 0, 255).astype(np.uint8)
    hsv[:, :, 2] = np.clip(np.rint(maximum * 255.0), 0, 255).astype(np.uint8)
    return hsv


def srgb_to_cielab(rgb: np.ndarray) -> np.ndarray:
    """Convert 0--255 IEC 61966-2-1 sRGB to CIE 1976 Lab (D65/2 degree).

    The conversion decodes the sRGB transfer curve, applies the standard D65
    sRGB-to-XYZ matrix, and uses the D65 reference white (0.95047, 1, 1.08883).
    It accepts any array whose final dimension has length three.
    """

    values = np.asarray(rgb, dtype=np.float64)
    if values.ndim == 0 or values.shape[-1] != 3:
        raise ValueError("RGB input must have a final dimension of length three")
    if not np.all(np.isfinite(values)) or np.any(values < 0) or np.any(values > 255):
        raise ValueError("RGB input must contain finite values on the 0--255 scale")

    encoded = values / 255.0
    linear = np.where(
        encoded <= 0.04045,
        encoded / 12.92,
        ((encoded + 0.055) / 1.055) ** 2.4,
    )
    matrix = np.asarray(
        [
            [0.4124564, 0.3575761, 0.1804375],
            [0.2126729, 0.7151522, 0.0721750],
            [0.0193339, 0.1191920, 0.9503041],
        ],
        dtype=np.float64,
    )
    xyz = linear @ matrix.T
    xyz /= np.asarray([0.95047, 1.0, 1.08883], dtype=np.float64)
    delta = 6.0 / 29.0
    transformed = np.where(
        xyz > delta**3,
        np.cbrt(xyz),
        xyz / (3.0 * delta**2) + 4.0 / 29.0,
    )
    x, y, z = transformed[..., 0], transformed[..., 1], transformed[..., 2]
    return np.stack((116.0 * y - 16.0, 500.0 * (x - y), 200.0 * (y - z)), axis=-1)


def estimate_joint_lab_peak(
    rgb_pixels: np.ndarray,
    *,
    weights: Optional[np.ndarray] = None,
    bin_width: float = PEAK_LAB_BIN_WIDTH,
    delta_e_radius: float = PEAK_DELTA_E_RADIUS,
    second_peak_min_delta_e: float = PEAK_SECOND_MIN_DELTA_E,
    multimodal_threshold: float = PEAK_MULTIMODAL_THRESHOLD,
) -> Dict[str, Any]:
    """Estimate the densest *joint* colour region in CIELAB space.

    Lab pixels are assigned to fixed 5-unit bins.  Bin weights are smoothed
    with a deterministic separable ``[1, 2, 1]`` kernel in three dimensions.
    The highest-scoring occupied bin seeds an explicit CIE76 DeltaE ball.  Its
    alpha-weighted Lab centroid is used to select its nearest observed pixel,
    so reported sRGB and Lab describe one internally consistent colour.
    A second peak must be at least ``second_peak_min_delta_e`` from the first,
    preventing adjacent bins in one broad mode from masquerading as two modes.
    No channel is trimmed or optimized independently.
    """

    pixels = np.asarray(rgb_pixels, dtype=np.float64)
    if pixels.ndim != 2 or pixels.shape[1] != 3 or len(pixels) == 0:
        raise ValueError("rgb_pixels must be a non-empty n-by-3 array")
    if not np.all(np.isfinite(pixels)) or np.any(pixels < 0) or np.any(pixels > 255):
        raise ValueError("rgb_pixels must be finite values on the 0--255 scale")
    if bin_width <= 0 or delta_e_radius <= 0 or second_peak_min_delta_e <= 0:
        raise ValueError("Lab bin width and DeltaE distances must be positive")
    if not (0 <= multimodal_threshold <= 1):
        raise ValueError("multimodal_threshold must be between zero and one")

    if weights is None:
        pixel_weights = np.ones(len(pixels), dtype=np.float64)
    else:
        pixel_weights = np.asarray(weights, dtype=np.float64)
        if pixel_weights.shape != (len(pixels),):
            raise ValueError("weights must contain one value per RGB pixel")
        if not np.all(np.isfinite(pixel_weights)) or np.any(pixel_weights < 0):
            raise ValueError("weights must be finite and non-negative")
        if not np.any(pixel_weights > 0):
            pixel_weights = np.ones(len(pixels), dtype=np.float64)

    lab = srgb_to_cielab(pixels)
    origin = np.asarray([0.0, -128.0, -128.0], dtype=np.float64)
    upper = np.asarray([100.0, 128.0, 128.0], dtype=np.float64)
    shape = tuple((np.floor((upper - origin) / bin_width).astype(int) + 1).tolist())
    coordinates = np.floor((lab - origin) / bin_width).astype(int)
    coordinates = np.clip(coordinates, 0, np.asarray(shape) - 1)
    packed = np.ravel_multi_index(coordinates.T, shape)
    occupied, inverse = np.unique(packed, return_inverse=True)
    occupied_weights = np.bincount(inverse, weights=pixel_weights)
    occupied_counts = np.bincount(inverse)
    occupied_coordinates = np.column_stack(np.unravel_index(occupied, shape))

    histogram = np.zeros(shape, dtype=np.float64)
    histogram[tuple(occupied_coordinates.T)] = occupied_weights
    smoothed_scores = np.zeros(len(occupied), dtype=np.float64)
    shape_array = np.asarray(shape)
    for offset_l in (-1, 0, 1):
        for offset_a in (-1, 0, 1):
            for offset_b in (-1, 0, 1):
                offset = np.asarray([offset_l, offset_a, offset_b])
                neighbours = occupied_coordinates + offset
                valid = np.all((neighbours >= 0) & (neighbours < shape_array), axis=1)
                kernel_weight = (
                    (2 if offset_l == 0 else 1)
                    * (2 if offset_a == 0 else 1)
                    * (2 if offset_b == 0 else 1)
                )
                smoothed_scores[valid] += kernel_weight * histogram[
                    tuple(neighbours[valid].T)
                ]

    # Highest smoothed density, then unsmoothed density, then fixed packed-bin
    # order.  The final tie-break makes the result independent of input order.
    order = np.lexsort((occupied, -occupied_counts, -occupied_weights, -smoothed_scores))
    primary_position = int(order[0])
    bin_centres = origin + (occupied_coordinates.astype(np.float64) + 0.5) * bin_width
    bin_centres = np.clip(bin_centres, origin, upper)
    primary_centre = bin_centres[primary_position]
    distance_to_bin_peak = np.linalg.norm(lab - primary_centre, axis=1)
    selected = distance_to_bin_peak <= delta_e_radius
    if not np.any(selected):  # Defensive: the occupied peak bin must intersect its own ball.
        selected = inverse == primary_position

    selected_weights = pixel_weights[selected]
    selected_lab = lab[selected]
    selected_rgb = pixels[selected]
    lab_centroid = np.average(selected_lab, axis=0, weights=selected_weights)
    centroid_distances = np.linalg.norm(selected_lab - lab_centroid, axis=1)
    nearest = np.flatnonzero(
        np.isclose(centroid_distances, np.min(centroid_distances), rtol=0, atol=1e-12)
    )
    if len(nearest) > 1:
        tie_rgb = selected_rgb[nearest]
        tie_order = np.lexsort((tie_rgb[:, 2], tie_rgb[:, 1], tie_rgb[:, 0]))
        representative_index = int(nearest[int(tie_order[0])])
    else:
        representative_index = int(nearest[0])
    # Report one observed sRGB/Lab pair nearest to the Lab centroid.  Averaging gamma-encoded sRGB
    # independently would produce an RGB triplet inconsistent with peak Lab.
    representative_rgb = selected_rgb[representative_index]
    representative_lab = selected_lab[representative_index]
    selected_distances = np.linalg.norm(selected_lab - representative_lab, axis=1)
    dispersion = math.sqrt(
        float(np.average(selected_distances**2, weights=selected_weights))
    )

    separated = np.linalg.norm(bin_centres - primary_centre, axis=1) >= second_peak_min_delta_e
    if np.any(separated):
        second_score = float(np.max(smoothed_scores[separated]))
        second_peak_ratio = second_score / float(smoothed_scores[primary_position])
    else:
        second_peak_ratio = 0.0
    second_peak_ratio = float(np.clip(second_peak_ratio, 0.0, 1.0))

    result: Dict[str, Any] = {
        "peak_pixels": int(np.count_nonzero(selected)),
        "peak_fraction": float(np.count_nonzero(selected) / len(pixels)),
        "peak_dispersion_deltaE": dispersion,
        "second_peak_ratio": second_peak_ratio,
        "multimodal_colour": bool(second_peak_ratio >= multimodal_threshold),
        "peak_method": "smoothed_fixed_bin_joint_cielab_d65_deltaE76_centroid_nearest_observed",
        "peak_lab_bin_width": float(bin_width),
        "peak_delta_e_radius": float(delta_e_radius),
        "peak_second_min_delta_e": float(second_peak_min_delta_e),
        "peak_multimodal_threshold": float(multimodal_threshold),
        "peak_low_fraction_threshold": PEAK_LOW_FRACTION_THRESHOLD,
    }
    for index, channel in enumerate("RGB"):
        result["peak_%s" % channel] = float(np.clip(representative_rgb[index], 0, 255))
    for index, channel in enumerate(("L", "a", "b")):
        result["peak_%s" % channel] = float(representative_lab[index])
    return result


def _empty_joint_lab_peak_metrics() -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "peak_pixels": 0,
        "peak_fraction": 0.0,
        "peak_dispersion_deltaE": math.nan,
        "second_peak_ratio": math.nan,
        "multimodal_colour": False,
        "peak_method": "smoothed_fixed_bin_joint_cielab_d65_deltaE76_centroid_nearest_observed",
        "peak_lab_bin_width": PEAK_LAB_BIN_WIDTH,
        "peak_delta_e_radius": PEAK_DELTA_E_RADIUS,
        "peak_second_min_delta_e": PEAK_SECOND_MIN_DELTA_E,
        "peak_multimodal_threshold": PEAK_MULTIMODAL_THRESHOLD,
        "peak_low_fraction_threshold": PEAK_LOW_FRACTION_THRESHOLD,
    }
    for channel in ("R", "G", "B", "L", "a", "b"):
        result["peak_%s" % channel] = math.nan
    return result


def _prefixed_peak_metrics(metrics: Mapping[str, Any], prefix: str) -> Dict[str, Any]:
    keys = [
        *("peak_%s" % channel for channel in ("R", "G", "B", "L", "a", "b")),
        "peak_pixels",
        "peak_fraction",
        "peak_dispersion_deltaE",
        "second_peak_ratio",
        "multimodal_colour",
    ]
    return {"%s_%s" % (prefix, key): metrics[key] for key in keys}


def mask_component_summary(
    mask: np.ndarray, *, large_component_min_fraction: float = MASK_LARGE_COMPONENT_MIN_FRACTION
) -> Dict[str, Any]:
    """Summarize 8-connected mask components with row-run union/find labeling."""

    binary = np.asarray(mask, dtype=bool)
    if binary.ndim != 2:
        raise ValueError("mask must be a two-dimensional array")
    if not (0 < large_component_min_fraction <= 1):
        raise ValueError("large_component_min_fraction must be in (0, 1]")
    mask_pixels = int(np.count_nonzero(binary))
    if mask_pixels == 0:
        return {
            "mask_component_count": 0,
            "mask_large_component_count": 0,
            "mask_largest_component_fraction": 0.0,
            "possible_multiple_flowers": False,
        }

    parents: List[int] = []
    sizes: List[int] = []

    def find(label: int) -> int:
        root = label
        while parents[root] != root:
            root = parents[root]
        while parents[label] != label:
            next_label = parents[label]
            parents[label] = root
            label = next_label
        return root

    def union(first: int, second: int) -> None:
        root_first, root_second = find(first), find(second)
        if root_first == root_second:
            return
        if sizes[root_first] < sizes[root_second]:
            root_first, root_second = root_second, root_first
        parents[root_second] = root_first
        sizes[root_first] += sizes[root_second]

    previous_runs: List[Tuple[int, int, int]] = []
    for row in binary:
        padded = np.pad(row.astype(np.int8), (1, 1))
        changes = np.diff(padded)
        starts = np.flatnonzero(changes == 1)
        ends = np.flatnonzero(changes == -1) - 1
        current_runs: List[Tuple[int, int, int]] = []
        previous_start = 0
        for start, end in zip(starts.tolist(), ends.tolist()):
            label = len(parents)
            parents.append(label)
            sizes.append(end - start + 1)
            while (
                previous_start < len(previous_runs)
                and previous_runs[previous_start][1] < start - 1
            ):
                previous_start += 1
            candidate = previous_start
            while candidate < len(previous_runs) and previous_runs[candidate][0] <= end + 1:
                union(label, previous_runs[candidate][2])
                candidate += 1
            current_runs.append((start, end, label))
        previous_runs = current_runs

    component_sizes: Dict[int, int] = {}
    for label, size in enumerate(sizes):
        if parents[label] == label:
            component_sizes[label] = size
    ordered_sizes = sorted(component_sizes.values(), reverse=True)
    large_minimum = large_component_min_fraction * mask_pixels
    large_count = sum(size >= large_minimum for size in ordered_sizes)
    return {
        "mask_component_count": len(ordered_sizes),
        "mask_large_component_count": large_count,
        "mask_largest_component_fraction": ordered_sizes[0] / mask_pixels,
        "possible_multiple_flowers": large_count >= 2,
    }


def _morphological_cleanup(mask: np.ndarray, kernel_size: int) -> np.ndarray:
    """Apply one binary opening followed by one closing using Pillow filters."""

    image = Image.fromarray(mask.astype(np.uint8) * 255, mode="L")
    # White foreground: MinFilter erodes and MaxFilter dilates.
    opened = image.filter(ImageFilter.MinFilter(kernel_size)).filter(
        ImageFilter.MaxFilter(kernel_size)
    )
    closed = opened.filter(ImageFilter.MaxFilter(kernel_size)).filter(
        ImageFilter.MinFilter(kernel_size)
    )
    return np.asarray(closed, dtype=np.uint8) > 0


def make_petal_mask(
    rgb: np.ndarray,
    alpha: np.ndarray,
    *,
    alpha_threshold: int = 128,
    pink_hue_low: int = 90,
    pink_hue_high: int = 179,
    red_hue_high: int = 10,
    colour_saturation_min: int = 40,
    colour_value_min: int = 40,
    white_saturation_max: int = 70,
    white_value_min: int = 150,
    morph_kernel: int = 3,
) -> Tuple[np.ndarray, str, int]:
    """Create an alpha-bounded blue-violet/pink/white petal candidate mask.

    Some source flowers are rendered cyan or blue by camera white balance.
    Hue 90--124 is therefore retained inside the already isolated alpha
    foreground; green foliage remains below this interval.
    """

    if rgb.shape[:2] != alpha.shape:
        raise ValueError("RGB and alpha dimensions differ")
    if not (0 <= alpha_threshold <= 255):
        raise ValueError("alpha_threshold must be between 0 and 255")
    visible = alpha >= alpha_threshold
    hsv = _rgb_to_opencv_hsv(rgb)
    hue, saturation, value = hsv[:, :, 0], hsv[:, :, 1], hsv[:, :, 2]

    pink = (
        (
            ((hue >= pink_hue_low) & (hue <= pink_hue_high))
            | (hue <= red_hue_high)
        )
        & (saturation >= colour_saturation_min)
        & (value >= colour_value_min)
    )
    white = (saturation <= white_saturation_max) & (value >= white_value_min)
    mask = visible & (pink | white)

    if morph_kernel < 0:
        raise ValueError("morph_kernel must be non-negative")
    if morph_kernel > 1:
        kernel_size = morph_kernel if morph_kernel % 2 == 1 else morph_kernel + 1
        mask = _morphological_cleanup(mask, kernel_size) & visible

    informative_alpha = bool(np.any(alpha < 255))
    mask_source = "alpha+hsv" if informative_alpha else "hsv"
    return mask, mask_source, int(np.count_nonzero(visible))


def extract_color_metrics(
    image_bytes: bytes,
    *,
    primary_statistic: str = "median",
    alpha_threshold: int = 128,
    morph_kernel: int = 3,
) -> Extraction:
    """Decode one image, segment petals, and calculate deterministic RGB metrics."""

    if primary_statistic not in {"mean", "median"}:
        raise ValueError("primary_statistic must be 'mean' or 'median'")
    (
        rgb,
        alpha,
        has_alpha,
        source_exif_present,
        camera_exif_present,
        white_balance_metadata_present,
        source_srgb_chunk_present,
        source_icc_profile_present,
    ) = _decode_image(image_bytes)
    mask, mask_source, visible_pixels = make_petal_mask(
        rgb,
        alpha,
        alpha_threshold=alpha_threshold,
        morph_kernel=morph_kernel,
    )
    height, width = mask.shape
    mask_pixels = int(np.count_nonzero(mask))
    total_pixels = int(mask.size)
    mask_fraction = mask_pixels / total_pixels if total_pixels else 0.0
    mask_fraction_visible = mask_pixels / visible_pixels if visible_pixels else 0.0

    metrics: Dict[str, Any] = {
        "image_width": width,
        "image_height": height,
        "has_alpha": bool(has_alpha),
        "alpha_threshold": alpha_threshold,
        "source_exif_present": source_exif_present,
        "camera_exif_present": camera_exif_present,
        "white_balance_metadata_present": white_balance_metadata_present,
        "source_srgb_chunk_present": source_srgb_chunk_present,
        "source_icc_profile_present": source_icc_profile_present,
        "source_colour_space": (
            "embedded_icc_profile_unconverted"
            if source_icc_profile_present
            else "sRGB" if source_srgb_chunk_present else "assumed_sRGB_unverified"
        ),
        "neutral_reference_available": False,
        "illumination_correction_status": "not_identifiable",
        "visible_pixels": visible_pixels,
        "mask_pixels": mask_pixels,
        "mask_fraction": mask_fraction,
        "mask_fraction_visible": mask_fraction_visible,
        "low_mask_coverage_threshold": LOW_MASK_COVERAGE_THRESHOLD,
        "mask_source": mask_source,
        "alpha_segmentation_source": "unknown_preprocessed_cutout",
        "morph_kernel_px": (
            0
            if morph_kernel <= 1
            else morph_kernel if morph_kernel % 2 == 1 else morph_kernel + 1
        ),
        "morph_kernel_short_side_fraction": (
            (morph_kernel if morph_kernel % 2 == 1 else morph_kernel + 1)
            / min(height, width)
            if morph_kernel > 1 and min(height, width) > 0
            else 0.0
        ),
        # The source cut-outs contain no independent grey/white reference.
        # Treating pale petal pixels as a reference would erase biological
        # colour variation, so raw decoded sRGB is deliberately retained.
        "white_balance_applied": False,
        "white_balance_method": "none",
        "white_balance_reliability": "unavailable",
        "white_balance_note": "postprocessed_srgb_cutout_no_neutral_reference",
        "color_statistic": primary_statistic,
        "primary_colour_method": "%s_hsv_mask_v2_1_compatible" % primary_statistic,
        "candidate_colour_methods": (
            "hsv_joint_lab_peak;hsv_exposure_filtered_joint_lab_peak;"
            "alpha_joint_lab_peak"
        ),
    }
    mask_components = mask_component_summary(mask)
    metrics.update(mask_components)
    visible_mask = alpha >= alpha_threshold
    alpha_large_fraction = min(
        1.0,
        max(0.01, 100.0 / visible_pixels) if visible_pixels else 1.0,
    )
    alpha_components = mask_component_summary(
        visible_mask, large_component_min_fraction=alpha_large_fraction
    )
    metrics["alpha_component_count"] = alpha_components["mask_component_count"]
    metrics["alpha_large_component_count"] = alpha_components[
        "mask_large_component_count"
    ]
    metrics["alpha_largest_component_fraction"] = alpha_components[
        "mask_largest_component_fraction"
    ]
    metrics["possible_multiple_components"] = bool(
        mask_components["possible_multiple_flowers"]
        or alpha_components["mask_large_component_count"] >= 2
    )
    hsv = _rgb_to_opencv_hsv(rgb)
    excluded_warm = (
        visible_mask
        & ~mask
        & (hsv[:, :, 0] >= 11)
        & (hsv[:, :, 0] <= 44)
        & (hsv[:, :, 1] > 70)
        & (hsv[:, :, 2] >= 150)
    )
    metrics["excluded_warm_fraction_visible"] = (
        float(np.count_nonzero(excluded_warm) / visible_pixels)
        if visible_pixels
        else 0.0
    )
    cool_blue = (
        visible_mask
        & (hsv[:, :, 0] >= 90)
        & (hsv[:, :, 0] <= 124)
        & (hsv[:, :, 1] > 70)
        & (hsv[:, :, 2] >= 150)
    )
    metrics["cool_blue_fraction_visible"] = (
        float(np.count_nonzero(cool_blue) / visible_pixels)
        if visible_pixels
        else 0.0
    )
    metrics["warm_sensitivity_mask_fraction_visible"] = (
        float(np.count_nonzero(mask | excluded_warm) / visible_pixels)
        if visible_pixels
        else 0.0
    )

    alpha_pixels = rgb[visible_mask].astype(np.float64)
    alpha_weights = alpha[visible_mask].astype(np.float64) / 255.0
    if len(alpha_pixels):
        alpha_peak = estimate_joint_lab_peak(alpha_pixels, weights=alpha_weights)
        metrics["alpha_all_channels_clipped_fraction"] = float(
            np.mean(np.all(alpha_pixels == 255, axis=1))
        )
        metrics["alpha_any_channel_clipped_fraction"] = float(
            np.mean(np.any(alpha_pixels == 255, axis=1))
        )
        metrics["alpha_dark_shadow_fraction"] = float(
            np.mean(srgb_to_cielab(alpha_pixels)[:, 0] < 25.0)
        )
    else:
        alpha_peak = _empty_joint_lab_peak_metrics()
        metrics["alpha_all_channels_clipped_fraction"] = 0.0
        metrics["alpha_any_channel_clipped_fraction"] = 0.0
        metrics["alpha_dark_shadow_fraction"] = 0.0
    metrics.update(_prefixed_peak_metrics(alpha_peak, "alpha"))
    for key in (
        "peak_method",
        "peak_lab_bin_width",
        "peak_delta_e_radius",
        "peak_second_min_delta_e",
        "peak_multimodal_threshold",
        "peak_low_fraction_threshold",
    ):
        metrics[key] = alpha_peak[key]
    metrics["alpha_hsv_peak_disagreement_threshold"] = (
        ALPHA_HSV_PEAK_DISAGREEMENT_THRESHOLD
    )
    flags: List[str] = []

    if mask_pixels == 0:
        for channel in "RGB":
            metrics["mean_%s" % channel] = math.nan
            metrics["median_%s" % channel] = math.nan
            metrics[channel] = math.nan
        metrics.update(
            _prefixed_peak_metrics(_empty_joint_lab_peak_metrics(), "hsv")
        )
        metrics.update(
            _prefixed_peak_metrics(
                _empty_joint_lab_peak_metrics(), "hsv_exposure_filtered"
            )
        )
        metrics["exposure_filtered_pixels"] = 0
        metrics["exposure_filtered_fraction"] = 0.0
        metrics["exposure_filtered_min_pixels"] = EXPOSURE_FILTERED_MIN_PIXELS
        metrics["exposure_filtered_min_fraction"] = EXPOSURE_FILTERED_MIN_FRACTION
        metrics["exposure_filtered_peak_available"] = False
        metrics["alpha_hsv_peak_deltaE"] = math.nan
        metrics["near_white_fraction"] = 0.0
        metrics["all_channels_clipped_fraction"] = 0.0
        metrics["any_channel_clipped_fraction"] = 0.0
        metrics["possible_overexposure"] = bool(
            metrics["alpha_all_channels_clipped_fraction"]
            > ALL_CHANNELS_CLIPPED_THRESHOLD
        )
        metrics["dark_shadow_fraction"] = 0.0
        metrics["qc_status"] = "no_mask"
        metrics["qc_flags"] = "no_mask"
        return Extraction(rgb=rgb, alpha=alpha, mask=mask, metrics=metrics)

    pixels = rgb[mask].astype(np.float64)
    weights = alpha[mask].astype(np.float64) / 255.0
    if not np.any(weights > 0):
        weights = np.ones_like(weights)
    means = np.average(pixels, axis=0, weights=weights)
    medians = np.median(pixels, axis=0)
    for index, channel in enumerate("RGB"):
        metrics["mean_%s" % channel] = float(means[index])
        metrics["median_%s" % channel] = float(medians[index])
        metrics[channel] = float((means if primary_statistic == "mean" else medians)[index])
    hsv_peak = estimate_joint_lab_peak(pixels, weights=weights)
    metrics.update(_prefixed_peak_metrics(hsv_peak, "hsv"))
    pixel_lab = srgb_to_cielab(pixels)
    exposure_retained = ~(
        np.all(pixels >= 250, axis=1) | (pixel_lab[:, 0] < 25.0)
    )
    retained_count = int(np.count_nonzero(exposure_retained))
    retained_fraction = retained_count / len(pixels)
    required_retained = max(
        EXPOSURE_FILTERED_MIN_PIXELS,
        int(math.ceil(EXPOSURE_FILTERED_MIN_FRACTION * len(pixels))),
    )
    metrics["exposure_filtered_pixels"] = retained_count
    metrics["exposure_filtered_fraction"] = retained_fraction
    metrics["exposure_filtered_min_pixels"] = EXPOSURE_FILTERED_MIN_PIXELS
    metrics["exposure_filtered_min_fraction"] = EXPOSURE_FILTERED_MIN_FRACTION
    metrics["exposure_filtered_peak_available"] = retained_count >= required_retained
    if metrics["exposure_filtered_peak_available"]:
        exposure_peak = estimate_joint_lab_peak(
            pixels[exposure_retained], weights=weights[exposure_retained]
        )
    else:
        exposure_peak = _empty_joint_lab_peak_metrics()
    metrics.update(_prefixed_peak_metrics(exposure_peak, "hsv_exposure_filtered"))
    metrics["alpha_hsv_peak_deltaE"] = float(
        math.sqrt(
            sum(
                (metrics["alpha_peak_%s" % channel] - metrics["hsv_peak_%s" % channel])
                ** 2
                for channel in ("L", "a", "b")
            )
        )
    )
    # This is a sensitivity heuristic, not proof of sensor clipping: genuine
    # white petals can also have rendered sRGB values at or above 250 in every
    # channel.
    metrics["near_white_fraction"] = float(
        np.mean(np.all(pixels >= 250, axis=1))
    )
    metrics["all_channels_clipped_fraction"] = float(
        np.mean(np.all(pixels == 255, axis=1))
    )
    metrics["any_channel_clipped_fraction"] = float(
        np.mean(np.any(pixels == 255, axis=1))
    )
    metrics["possible_overexposure"] = bool(
        max(
            metrics["all_channels_clipped_fraction"],
            metrics["alpha_all_channels_clipped_fraction"],
        )
        > ALL_CHANNELS_CLIPPED_THRESHOLD
    )
    metrics["dark_shadow_fraction"] = float(np.mean(pixel_lab[:, 0] < 25.0))

    border_pixels = int(
        np.count_nonzero(mask[0, :])
        + np.count_nonzero(mask[-1, :])
        + np.count_nonzero(mask[1:-1, 0])
        + np.count_nonzero(mask[1:-1, -1])
    )
    if mask_fraction_visible < 0.05:
        flags.append("low_mask_fraction")
    if mask_fraction_visible < LOW_MASK_COVERAGE_THRESHOLD:
        flags.append("low_mask_coverage")
    if not np.any(alpha < 255) and mask_fraction > 0.95:
        flags.append("high_mask_fraction")
    if border_pixels / mask_pixels > 0.10:
        flags.append("touches_border")
    if metrics["possible_multiple_components"]:
        flags.append("possible_multiple_components")
    if metrics["excluded_warm_fraction_visible"] > EXCLUDED_WARM_THRESHOLD:
        flags.append("possible_warm_cast_or_petal_exclusion")
    if metrics["cool_blue_fraction_visible"] > COOL_BLUE_THRESHOLD:
        flags.append("possible_cool_cast_or_blue_petals")
    if metrics["possible_overexposure"]:
        flags.append("possible_overexposure")
    if metrics["near_white_fraction"] > NEAR_WHITE_FRACTION_THRESHOLD:
        flags.append("high_near_white_fraction")
    if metrics["dark_shadow_fraction"] > DARK_SHADOW_THRESHOLD:
        flags.append("dark_shadow_dominated")
    if metrics["hsv_multimodal_colour"]:
        flags.append("hsv_multimodal_colour")
    if metrics["alpha_multimodal_colour"]:
        flags.append("alpha_multimodal_colour")
    if metrics["hsv_peak_fraction"] < PEAK_LOW_FRACTION_THRESHOLD:
        flags.append("low_hsv_peak_fraction")
    if metrics["alpha_peak_fraction"] < PEAK_LOW_FRACTION_THRESHOLD:
        flags.append("low_alpha_peak_fraction")
    if not metrics["exposure_filtered_peak_available"]:
        flags.append("insufficient_exposure_filtered_pixels")
    elif metrics["hsv_exposure_filtered_multimodal_colour"]:
        flags.append("hsv_exposure_filtered_multimodal_colour")
    if (
        metrics["alpha_hsv_peak_deltaE"]
        > ALPHA_HSV_PEAK_DISAGREEMENT_THRESHOLD
    ):
        flags.append("alpha_hsv_peak_disagreement")
    if source_icc_profile_present:
        flags.append("icc_profile_unconverted")
    elif not source_srgb_chunk_present:
        flags.append("assumed_srgb_unverified")
    metrics["qc_status"] = "manual_review_required" if flags else "ok"
    metrics["qc_flags"] = ";".join(flags)
    return Extraction(rgb=rgb, alpha=alpha, mask=mask, metrics=metrics)


def _safe_stem(observation_id: str, digest: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", observation_id).strip("._")
    return "%s_%s" % ((cleaned or "observation")[:100], digest[:10] or "noimage")


def _checker_composite(rgb: np.ndarray, alpha: np.ndarray) -> np.ndarray:
    """Composite RGBA content over a white/light-grey checkerboard."""

    height, width = alpha.shape
    tile = max(6, min(height, width) // 16)
    yy, xx = np.indices((height, width))
    light = ((xx // tile + yy // tile) % 2) == 0
    checker = np.empty((height, width, 3), dtype=np.float32)
    checker[light] = 255.0
    checker[~light] = 220.0
    fraction = alpha.astype(np.float32)[:, :, None] / 255.0
    return np.clip(rgb.astype(np.float32) * fraction + checker * (1.0 - fraction), 0, 255).astype(
        np.uint8
    )


def _display_size(image: Image.Image, max_height: int = 480) -> Tuple[int, int]:
    if image.height <= max_height:
        return image.width, image.height
    scale = max_height / image.height
    return max(1, int(round(image.width * scale))), max_height


def _swatch_colour(metrics: Mapping[str, Any], statistic: str) -> Optional[Tuple[int, int, int]]:
    values: List[int] = []
    for channel in "RGB":
        value = metrics.get("%s_%s" % (statistic, channel))
        if value is None or not math.isfinite(float(value)):
            return None
        values.append(int(round(float(value))))
    return tuple(values)  # type: ignore[return-value]


def _text_colour(background: Tuple[int, int, int]) -> Tuple[int, int, int]:
    luminance = 0.2126 * background[0] + 0.7152 * background[1] + 0.0722 * background[2]
    return (0, 0, 0) if luminance >= 150 else (255, 255, 255)


def _make_qc_panel(
    extraction: Extraction,
    original: Image.Image,
    mask_image: Image.Image,
    overlay: Image.Image,
) -> Image.Image:
    """Build an original/alpha/HSV-mask/overlay/four-estimator QC panel."""

    target_size = _display_size(original)
    original_view = original.resize(target_size, Image.Resampling.LANCZOS)
    alpha_threshold = int(extraction.metrics.get("alpha_threshold") or 128)
    alpha_image = Image.fromarray(
        (extraction.alpha >= alpha_threshold).astype(np.uint8) * 255, mode="L"
    )
    alpha_view = alpha_image.convert("RGB").resize(target_size, Image.Resampling.NEAREST)
    mask_view = mask_image.convert("RGB").resize(target_size, Image.Resampling.NEAREST)
    overlay_view = overlay.resize(target_size, Image.Resampling.LANCZOS)
    view_width, view_height = target_size
    swatch_width = max(180, min(view_width, 320))
    swatch = Image.new("RGB", (swatch_width, view_height), (205, 205, 205))
    swatch_draw = ImageDraw.Draw(swatch)
    boundaries = [int(round(view_height * index / 5.0)) for index in range(6)]
    for statistic, label, top, bottom in (
        ("mean", "HSV-mask mean", boundaries[0], boundaries[1]),
        ("median", "HSV-mask median", boundaries[1], boundaries[2]),
        ("hsv_peak", "HSV-mask Lab peak", boundaries[2], boundaries[3]),
        (
            "hsv_exposure_filtered_peak",
            "HSV exposure-filtered peak",
            boundaries[3],
            boundaries[4],
        ),
        ("alpha_peak", "Alpha Lab peak", boundaries[4], boundaries[5]),
    ):
        colour = _swatch_colour(extraction.metrics, statistic)
        fill = colour or (205, 205, 205)
        swatch_draw.rectangle((0, top, swatch_width, max(top, bottom - 1)), fill=fill)
        value_text = (
            "%d, %d, %d" % colour if colour is not None else "NA (no petal mask)"
        )
        swatch_draw.text(
            (8, top + 6),
            "%s\n%s" % (label, value_text),
            fill=_text_colour(fill),
        )

    header_height = 26
    gap = 6
    sections = [
        ("Original", original_view),
        ("Alpha foreground", alpha_view),
        ("HSV petal mask", mask_view),
        ("Overlay", overlay_view),
        ("Mean / median / joint peaks", swatch),
    ]
    panel_width = sum(section.width for _, section in sections) + gap * (len(sections) - 1)
    panel = Image.new("RGB", (panel_width, header_height + view_height), (255, 255, 255))
    draw = ImageDraw.Draw(panel)
    left = 0
    for label, section in sections:
        draw.text((left + 4, 6), label, fill=(0, 0, 0))
        panel.paste(section, (left, header_height))
        left += section.width + gap
    return panel


def write_qc_images(
    extraction: Extraction,
    *,
    observation_id: str,
    digest: str,
    qc_dir: Path,
    overwrite: bool,
) -> Tuple[str, str, str]:
    masks_dir = qc_dir / "masks"
    overlays_dir = qc_dir / "overlays"
    panels_dir = qc_dir / "panels"
    masks_dir.mkdir(parents=True, exist_ok=True)
    overlays_dir.mkdir(parents=True, exist_ok=True)
    panels_dir.mkdir(parents=True, exist_ok=True)
    stem = _safe_stem(observation_id, digest)
    mask_path = masks_dir / (stem + ".png")
    overlay_path = overlays_dir / (stem + ".png")
    panel_path = panels_dir / (stem + ".png")
    for path in (mask_path, overlay_path, panel_path):
        if path.exists() and not overwrite:
            raise FileExistsError("QC output already exists: %s" % path)

    mask_image = Image.fromarray(extraction.mask.astype(np.uint8) * 255, mode="L")
    mask_image.save(mask_path)
    base = _checker_composite(extraction.rgb, extraction.alpha)
    overlay = base.copy()
    if np.any(extraction.mask):
        red = np.zeros_like(overlay)
        red[:, :, 0] = 255
        overlay[extraction.mask] = (
            0.65 * overlay[extraction.mask] + 0.35 * red[extraction.mask]
        ).astype(np.uint8)
        mask_image = Image.fromarray(extraction.mask.astype(np.uint8) * 255, mode="L")
        dilated = np.asarray(mask_image.filter(ImageFilter.MaxFilter(3)), dtype=np.int16)
        eroded = np.asarray(mask_image.filter(ImageFilter.MinFilter(3)), dtype=np.int16)
        boundary = (dilated - eroded) > 0
        overlay[boundary] = np.array([255, 255, 0], dtype=np.uint8)
    original_image = Image.fromarray(base, mode="RGB")
    overlay_image = Image.fromarray(overlay, mode="RGB")
    overlay_image.save(overlay_path)
    panel = _make_qc_panel(extraction, original_image, mask_image, overlay_image)
    panel.save(panel_path)
    return str(mask_path), str(overlay_path), str(panel_path)


def _base_output(record: SourceRecord) -> Dict[str, Any]:
    return {
        "observation_id": record.observation_id,
        "photo_id": record.photo_id,
        "source_row": record.source_row,
        "source_sheet": record.source_sheet,
        "source_image": record.source_image,
        "source_path": (
            (record.source_image or record.file_path.name)
            if record.file_path is not None
            else (
                "%s#%s" % (record.workbook_path.name, record.zip_member)
                if record.workbook_path is not None and record.zip_member
                else record.workbook_path.name if record.workbook_path is not None else None
            )
        ),
        "url": _normalise_scalar(record.url),
        "date": _normalise_scalar(record.date),
        "latitude": _normalise_scalar(record.latitude),
        "longitude": _normalise_scalar(record.longitude),
        "legacy_R": _normalise_scalar(record.legacy_R),
        "legacy_G": _normalise_scalar(record.legacy_G),
        "legacy_B": _normalise_scalar(record.legacy_B),
    }


def _empty_metrics(status: str, flag: str, primary_statistic: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "image_width": None,
        "image_height": None,
        "has_alpha": None,
        "alpha_threshold": None,
        "source_exif_present": None,
        "camera_exif_present": None,
        "white_balance_metadata_present": None,
        "source_srgb_chunk_present": None,
        "source_icc_profile_present": None,
        "source_colour_space": None,
        "neutral_reference_available": False,
        "illumination_correction_status": "not_identifiable",
        "visible_pixels": 0,
        "mask_pixels": 0,
        "mask_fraction": 0.0,
        "mask_fraction_visible": 0.0,
        "low_mask_coverage_threshold": LOW_MASK_COVERAGE_THRESHOLD,
        "mask_source": None,
        "alpha_segmentation_source": None,
        "morph_kernel_px": None,
        "morph_kernel_short_side_fraction": None,
        "mask_component_count": 0,
        "mask_large_component_count": 0,
        "mask_largest_component_fraction": 0.0,
        "possible_multiple_flowers": False,
        "alpha_component_count": 0,
        "alpha_large_component_count": 0,
        "alpha_largest_component_fraction": 0.0,
        "possible_multiple_components": False,
        "excluded_warm_fraction_visible": 0.0,
        "cool_blue_fraction_visible": 0.0,
        "warm_sensitivity_mask_fraction_visible": 0.0,
        "near_white_fraction": 0.0,
        "all_channels_clipped_fraction": 0.0,
        "any_channel_clipped_fraction": 0.0,
        "alpha_all_channels_clipped_fraction": 0.0,
        "alpha_any_channel_clipped_fraction": 0.0,
        "possible_overexposure": False,
        "dark_shadow_fraction": 0.0,
        "alpha_dark_shadow_fraction": 0.0,
        "white_balance_applied": False,
        "white_balance_method": "none",
        "white_balance_reliability": "unavailable",
        "white_balance_note": "postprocessed_srgb_cutout_no_neutral_reference",
        "color_statistic": primary_statistic,
        "primary_colour_method": "%s_hsv_mask_v2_1_compatible" % primary_statistic,
        "candidate_colour_methods": (
            "hsv_joint_lab_peak;hsv_exposure_filtered_joint_lab_peak;"
            "alpha_joint_lab_peak"
        ),
        "extraction_version": EXTRACTION_VERSION,
        "qc_status": status,
        "qc_flags": flag,
        "qc_note": flag,
        "mask_path": None,
        "overlay_path": None,
        "qc_panel_path": None,
    }
    for channel in "RGB":
        out["mean_%s" % channel] = math.nan
        out["median_%s" % channel] = math.nan
        out[channel] = math.nan
    empty_peak = _empty_joint_lab_peak_metrics()
    out.update(_prefixed_peak_metrics(empty_peak, "hsv"))
    out.update(_prefixed_peak_metrics(empty_peak, "hsv_exposure_filtered"))
    out.update(_prefixed_peak_metrics(empty_peak, "alpha"))
    out["exposure_filtered_pixels"] = 0
    out["exposure_filtered_fraction"] = 0.0
    out["exposure_filtered_min_pixels"] = EXPOSURE_FILTERED_MIN_PIXELS
    out["exposure_filtered_min_fraction"] = EXPOSURE_FILTERED_MIN_FRACTION
    out["exposure_filtered_peak_available"] = False
    for key in (
        "peak_method",
        "peak_lab_bin_width",
        "peak_delta_e_radius",
        "peak_second_min_delta_e",
        "peak_multimodal_threshold",
        "peak_low_fraction_threshold",
    ):
        out[key] = empty_peak[key]
    out["alpha_hsv_peak_deltaE"] = math.nan
    out["alpha_hsv_peak_disagreement_threshold"] = (
        ALPHA_HSV_PEAK_DISAGREEMENT_THRESHOLD
    )
    return out


def _record_image_bytes(
    record: SourceRecord, archives: MutableMapping[Path, zipfile.ZipFile]
) -> bytes:
    if record.file_path is not None:
        return record.file_path.read_bytes()
    if record.workbook_path is not None and record.zip_member:
        archive = archives.get(record.workbook_path)
        if archive is None:
            archive = zipfile.ZipFile(record.workbook_path)
            archives[record.workbook_path] = archive
        try:
            return archive.read(record.zip_member)
        except KeyError as exc:
            raise FileNotFoundError("Workbook media is missing: %s" % record.zip_member) from exc
    raise FileNotFoundError("No image is mapped to this observation")


def _hash_rank(observation_id: str) -> Tuple[bytes, str]:
    return hashlib.sha256(observation_id.encode("utf-8")).digest(), observation_id


def _bin_by_cutpoints(value: float, lower: float, upper: float) -> str:
    if value <= lower:
        return "low"
    if value >= upper:
        return "high"
    return "mid"


def _finite_number(value: Any) -> Optional[float]:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _legacy_state(row: Mapping[str, Any]) -> str:
    values = [row.get("legacy_%s" % channel) for channel in "RGB"]
    if all(_finite_number(value) is not None for value in values):
        return "complete"
    missing = [
        value is None
        or (isinstance(value, str) and not value.strip())
        or (isinstance(value, (float, np.floating)) and math.isnan(float(value)))
        for value in values
    ]
    return "all_missing" if all(missing) else "partial"


def _add_legacy_deltas(row: MutableMapping[str, Any]) -> None:
    """Add signed v2-minus-legacy channel differences and RGB distance."""

    deltas: List[float] = []
    for channel in "RGB":
        current = _finite_number(row.get(channel))
        legacy = _finite_number(row.get("legacy_%s" % channel))
        delta = current - legacy if current is not None and legacy is not None else math.nan
        row["delta_%s_vs_legacy" % channel] = delta
        deltas.append(delta)
    row["delta_rgb_euclidean_vs_legacy"] = (
        math.sqrt(sum(delta * delta for delta in deltas))
        if all(math.isfinite(delta) for delta in deltas)
        else math.nan
    )


def _select_qc_indexes(
    records: Sequence[SourceRecord],
    rows: Sequence[Mapping[str, Any]],
    sample_size: Optional[int],
) -> Tuple[set[int], Dict[int, List[str]]]:
    """Select mandatory legacy/QC cases plus a separate stratified sample."""

    if sample_size is None or sample_size >= len(rows):
        return set(range(len(rows))), {
            index: ["all_observations_requested"] for index in range(len(rows))
        }

    reasons: Dict[int, List[str]] = {}

    def require(index: int, reason: str) -> None:
        reasons.setdefault(index, [])
        if reason not in reasons[index]:
            reasons[index].append(reason)

    for index, row in enumerate(rows):
        if (
            row.get("qc_status") != "ok"
            or bool(row.get("qc_flags"))
            or bool(row.get("duplicate_image_sha256"))
        ):
            require(index, "non_ok_or_duplicate")

    legacy_complete_indexes = [
        index for index, row in enumerate(rows) if _legacy_state(row) == "complete"
    ]
    if legacy_complete_indexes:
        for index, row in enumerate(rows):
            if _legacy_state(row) == "all_missing":
                require(index, "legacy_rgb_all_missing")
        comparable = [
            index
            for index in legacy_complete_indexes
            if _finite_number(rows[index].get("delta_rgb_euclidean_vs_legacy")) is not None
        ]
        comparable.sort(
            key=lambda index: (
                -float(rows[index]["delta_rgb_euclidean_vs_legacy"]),
                _hash_rank(records[index].observation_id),
            )
        )
        for index in comparable[:20]:
            require(index, "top20_legacy_rgb_difference")

    for metric, reason in (
        ("excluded_warm_fraction_visible", "top20_excluded_warm_fraction"),
        ("cool_blue_fraction_visible", "top20_cool_blue_fraction"),
    ):
        ranked = [
            index
            for index, row in enumerate(rows)
            if (_finite_number(row.get(metric)) or 0.0) > 0
        ]
        ranked.sort(
            key=lambda index: (
                -float(rows[index][metric]),
                _hash_rank(records[index].observation_id),
            )
        )
        for index in ranked[:20]:
            require(index, reason)

    mandatory = set(reasons)
    # The requested stratified count is additional to all mandatory cases.
    target = min(len(rows), len(mandatory) + sample_size)
    candidates = [index for index in range(len(rows)) if index not in mandatory]
    if len(mandatory) >= target or not candidates:
        return mandatory, reasons

    usable: List[int] = []
    brightness: Dict[int, float] = {}
    saturation: Dict[int, float] = {}
    mask_fraction: Dict[int, float] = {}
    for index in candidates:
        row = rows[index]
        try:
            rgb = [float(row["median_%s" % channel]) for channel in "RGB"]
            fraction = float(row["mask_fraction"])
        except (KeyError, TypeError, ValueError):
            continue
        if not all(math.isfinite(value) for value in rgb + [fraction]):
            continue
        maximum, minimum = max(rgb), min(rgb)
        brightness[index] = 0.2126 * rgb[0] + 0.7152 * rgb[1] + 0.0722 * rgb[2]
        saturation[index] = (maximum - minimum) / maximum if maximum > 0 else 0.0
        mask_fraction[index] = fraction
        usable.append(index)

    strata: Dict[Tuple[str, str, str], List[int]] = {}
    if usable:
        brightness_values = np.asarray([brightness[index] for index in usable], dtype=float)
        fraction_values = np.asarray([mask_fraction[index] for index in usable], dtype=float)
        b_low, b_high = np.quantile(brightness_values, [1.0 / 3.0, 2.0 / 3.0])
        f_low, f_high = np.quantile(fraction_values, [1.0 / 3.0, 2.0 / 3.0])
        brightness_median = float(np.median(brightness_values))
        for index in usable:
            if saturation[index] <= 0.18 and brightness[index] >= 160:
                colour_class = "white"
            elif brightness[index] <= brightness_median:
                colour_class = "deep_pink"
            else:
                colour_class = "light_pink"
            key = (
                colour_class,
                _bin_by_cutpoints(brightness[index], float(b_low), float(b_high)),
                _bin_by_cutpoints(mask_fraction[index], float(f_low), float(f_high)),
            )
            strata.setdefault(key, []).append(index)

    unusable = [index for index in candidates if index not in set(usable)]
    if unusable:
        strata[("other", "other", "other")] = unusable
    for indexes in strata.values():
        indexes.sort(key=lambda index: _hash_rank(records[index].observation_id))

    selected = set(mandatory)
    stratum_keys = sorted(strata)
    while len(selected) < target and stratum_keys:
        remaining_keys: List[Tuple[str, str, str]] = []
        for key in stratum_keys:
            indexes = strata[key]
            if indexes and len(selected) < target:
                index = indexes.pop(0)
                selected.add(index)
                require(index, "stratified_colour_brightness_mask")
            if indexes:
                remaining_keys.append(key)
        stratum_keys = remaining_keys
    return selected, reasons


def _build_qc_note(
    row: Mapping[str, Any], sampled: bool, selection_reasons: Sequence[str] = ()
) -> str:
    notes: List[str] = []
    status = str(row.get("qc_status") or "")
    if status == "no_mask":
        notes.append("No pixels met the alpha-bounded HSV petal criteria.")
    elif status == "missing_image":
        notes.append("No image was mapped to this observation.")
    elif status == "unreadable_image":
        notes.append("The image could not be decoded.")
    flags = [flag for flag in str(row.get("qc_flags") or "").split(";") if flag]
    if flags and flags != ["no_mask"]:
        notes.append("Manual review flags: %s." % ", ".join(flags))
    if sampled:
        if selection_reasons:
            notes.append("QC selection: %s." % ", ".join(selection_reasons))
        notes.append(
            "QC panel selected for manual visual review, including possible multiple flowers."
        )
    return " ".join(notes)


def process_records(
    records: Sequence[SourceRecord],
    *,
    qc_dir: Path,
    primary_statistic: str = "median",
    alpha_threshold: int = 128,
    morph_kernel: int = 3,
    qc_sample_size: Optional[int] = None,
    processed_at: Optional[str] = None,
    overwrite: bool = False,
) -> List[Dict[str, Any]]:
    """Process records, then create panels for problems and a stratified sample."""

    if qc_sample_size is not None and qc_sample_size < 0:
        raise ValueError("qc_sample_size must be non-negative")
    run_timestamp = processed_at or processing_timestamp()

    results: List[Dict[str, Any]] = []
    workbook_archives: Dict[Path, zipfile.ZipFile] = {}
    try:
        for record in records:
            row = _base_output(record)
            row["extraction_version"] = EXTRACTION_VERSION
            row["processed_at"] = run_timestamp
            row["qc_sampled"] = False
            row["mask_path"] = None
            row["overlay_path"] = None
            row["qc_panel_path"] = None
            try:
                image_bytes = _record_image_bytes(record, workbook_archives)
            except FileNotFoundError as exc:
                row.update(_empty_metrics("missing_image", str(exc), primary_statistic))
                row["image_sha256"] = None
                results.append(row)
                continue
            except OSError as exc:
                row.update(_empty_metrics("unreadable_image", str(exc), primary_statistic))
                row["image_sha256"] = None
                results.append(row)
                continue

            digest = sha256_bytes(image_bytes)
            row["image_sha256"] = digest
            try:
                extraction = extract_color_metrics(
                    image_bytes,
                    primary_statistic=primary_statistic,
                    alpha_threshold=alpha_threshold,
                    morph_kernel=morph_kernel,
                )
                row.update(extraction.metrics)
            except ValueError as exc:
                row.update(
                    _empty_metrics("unreadable_image", str(exc), primary_statistic)
                )
            results.append(row)
    finally:
        for archive in workbook_archives.values():
            archive.close()

    for row in results:
        _add_legacy_deltas(row)

    counts: Dict[str, int] = {}
    for row in results:
        digest = row.get("image_sha256")
        if digest:
            counts[digest] = counts.get(digest, 0) + 1
    for row in results:
        duplicate = bool(row.get("image_sha256") and counts[row["image_sha256"]] > 1)
        row["duplicate_image_sha256"] = duplicate
        if duplicate:
            flags = [flag for flag in str(row.get("qc_flags") or "").split(";") if flag]
            if "duplicate_image" not in flags:
                flags.append("duplicate_image")
            row["qc_flags"] = ";".join(flags)
            if row.get("qc_status") == "ok":
                row["qc_status"] = "manual_review_required"

    qc_indexes, qc_reasons = _select_qc_indexes(records, results, qc_sample_size)
    workbook_archives = {}
    try:
        for index in sorted(qc_indexes):
            row = results[index]
            row["qc_sampled"] = True
            try:
                image_bytes = _record_image_bytes(records[index], workbook_archives)
                extraction = extract_color_metrics(
                    image_bytes,
                    primary_statistic=primary_statistic,
                    alpha_threshold=alpha_threshold,
                    morph_kernel=morph_kernel,
                )
                mask_path, overlay_path, panel_path = write_qc_images(
                    extraction,
                    observation_id=records[index].observation_id,
                    digest=str(row.get("image_sha256") or sha256_bytes(image_bytes)),
                    qc_dir=qc_dir,
                    overwrite=overwrite,
                )
                row["mask_path"] = mask_path
                row["overlay_path"] = overlay_path
                row["qc_panel_path"] = panel_path
            except (FileNotFoundError, OSError, ValueError):
                # The row remains selected and retains its explicit failure status,
                # but a panel cannot be rendered without a decodable image.
                pass
    finally:
        for archive in workbook_archives.values():
            archive.close()

    for index, row in enumerate(results):
        row["qc_note"] = _build_qc_note(
            row, index in qc_indexes, qc_reasons.get(index, ())
        )
    return results


def _blankable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (float, np.floating)) and math.isnan(float(value)):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_results(
    rows: Sequence[Mapping[str, Any]], output_path: Path, *, overwrite: bool = False
) -> None:
    """Atomically write CSV or XLSX output; never overwrite by default."""

    suffix = output_path.suffix.casefold()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Output must end in .csv or .xlsx")
    if output_path.exists() and not overwrite:
        raise FileExistsError("Output already exists: %s" % output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(
        prefix=output_path.stem + ".",
        suffix=output_path.suffix,
        dir=output_path.parent,
        delete=False,
    )
    temp_path = Path(handle.name)
    handle.close()
    try:
        if suffix == ".csv":
            with temp_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
                writer.writeheader()
                for row in rows:
                    writer.writerow({key: _blankable(row.get(key)) for key in OUTPUT_COLUMNS})
        else:
            try:
                from openpyxl import Workbook
            except ImportError as exc:
                raise RuntimeError("XLSX output requires openpyxl; install .[excel]") from exc
            workbook = Workbook(write_only=True)
            worksheet = workbook.create_sheet("color_v2")
            worksheet.append(OUTPUT_COLUMNS)
            for row in rows:
                worksheet.append([_blankable(row.get(key)) for key in OUTPUT_COLUMNS])
            workbook.save(temp_path)
        os.replace(temp_path, output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Extract alpha-bounded HSV petal colours with reproducible IDs and QC outputs."
        )
    )
    inputs = parser.add_mutually_exclusive_group(required=True)
    inputs.add_argument("--input-dir", type=Path, help="Directory of image files")
    inputs.add_argument(
        "--input-workbook", type=Path, help="XLSX workbook containing in-cell images"
    )
    parser.add_argument("--output", required=True, type=Path, help="New .csv or .xlsx output")
    parser.add_argument("--qc-dir", required=True, type=Path, help="Directory for masks and overlays")
    parser.add_argument("--sheet", help="Workbook sheet name (default: first sheet)")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--image-column", default="petal", help="Header name or Excel column")
    parser.add_argument("--id-column", default="observation_id", help="Header name or Excel column")
    parser.add_argument(
        "--primary-statistic", choices=("mean", "median"), default="median"
    )
    parser.add_argument("--alpha-threshold", type=int, default=128)
    parser.add_argument("--morph-kernel", type=int, default=3)
    parser.add_argument(
        "--qc-sample-size",
        type=int,
        default=20,
        help="Deterministic QC mask/overlay sample (minimum 20; default: 20)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow replacing the requested output and matching QC files",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    if args.header_row < 1:
        raise SystemExit("--header-row must be at least 1")
    if not (0 <= args.alpha_threshold <= 255):
        raise SystemExit("--alpha-threshold must be between 0 and 255")
    if args.morph_kernel < 0:
        raise SystemExit("--morph-kernel must be non-negative")
    if args.qc_sample_size < 20:
        raise SystemExit("--qc-sample-size must be at least 20")

    input_path = args.input_dir or args.input_workbook
    assert input_path is not None
    if os.path.abspath(str(input_path)) == os.path.abspath(str(args.output)):
        raise SystemExit("Refusing to overwrite the input path")
    if args.output.exists() and not args.overwrite:
        raise SystemExit("Output already exists (use --overwrite for a derived file): %s" % args.output)

    try:
        if args.input_dir is not None:
            records = collect_directory_records(args.input_dir)
        else:
            records = collect_workbook_records(
                args.input_workbook,
                sheet_name=args.sheet,
                header_row=args.header_row,
                image_column=args.image_column,
                id_column=args.id_column,
            )
        if not records:
            raise ValueError("No image observations found")
        rows = process_records(
            records,
            qc_dir=args.qc_dir,
            primary_statistic=args.primary_statistic,
            alpha_threshold=args.alpha_threshold,
            morph_kernel=args.morph_kernel,
            qc_sample_size=args.qc_sample_size,
            overwrite=args.overwrite,
        )
        write_results(rows, args.output, overwrite=args.overwrite)
    except (FileExistsError, OSError, RuntimeError, ValueError, zipfile.BadZipFile) as exc:
        raise SystemExit(str(exc)) from exc

    status_counts: Dict[str, int] = {}
    for row in rows:
        status = str(row.get("qc_status") or "unknown")
        status_counts[status] = status_counts.get(status, 0) + 1
    summary = ", ".join("%s=%d" % item for item in sorted(status_counts.items()))
    print("Wrote %d observations to %s (%s)" % (len(rows), args.output, summary))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
