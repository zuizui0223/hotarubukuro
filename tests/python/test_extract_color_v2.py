import csv
import io
import math
import zipfile
from dataclasses import replace
from pathlib import Path

import numpy as np
import pytest
from PIL import Image

from scripts.extract_color_v2 import (
    collect_directory_records,
    collect_workbook_records,
    estimate_joint_lab_peak,
    extract_color_metrics,
    main,
    mask_component_summary,
    process_records,
    srgb_to_cielab,
    workbook_image_map,
    write_results,
)


def test_python_and_r_share_one_srgb_cielab_d65_fixture() -> None:
    fixture = np.genfromtxt(
        Path(__file__).parents[1] / "fixtures" / "srgb_cielab_d65.csv",
        delimiter=",",
        names=True,
    )
    rgb = np.column_stack([fixture[name] for name in ("R", "G", "B")])
    expected = np.column_stack([fixture[name] for name in ("L", "a", "b")])
    np.testing.assert_allclose(srgb_to_cielab(rgb), expected, rtol=0, atol=1e-10)


def png_bytes(array, mode="RGBA", exif=None):
    stream = io.BytesIO()
    if exif is None:
        exif = Image.Exif()
    exif[40961] = 1  # Explicit sRGB ColorSpace for provenance tests.
    save_options = {"exif": exif}
    Image.fromarray(array, mode=mode).save(stream, format="PNG", **save_options)
    return stream.getvalue()


def transparent_flower(colour=(210, 70, 170), size=40):
    # Deliberately give transparent pixels a petal-like RGB value.  Alpha must
    # exclude them from both the mask and summary statistics.
    rgba = np.zeros((size, size, 4), dtype=np.uint8)
    rgba[:, :, :3] = (255, 0, 255)
    rgba[10:30, 10:30, :3] = colour
    rgba[10:30, 10:30, 3] = 255
    return rgba


def test_alpha_bounds_hsv_mask_and_rgb_statistics():
    colour = (210, 70, 170)
    result = extract_color_metrics(
        png_bytes(transparent_flower(colour)), morph_kernel=3
    )

    assert result.metrics["mask_source"] == "alpha+hsv"
    assert result.metrics["mask_pixels"] == 400
    assert result.metrics["visible_pixels"] == 400
    assert result.metrics["qc_status"] == "ok"
    assert tuple(result.metrics[channel] for channel in "RGB") == colour
    assert tuple(result.metrics[f"mean_{channel}"] for channel in "RGB") == colour
    assert tuple(result.metrics[f"median_{channel}"] for channel in "RGB") == colour
    assert not result.mask[0, 0]
    assert result.metrics["white_balance_applied"] is False
    assert result.metrics["white_balance_method"] == "none"
    assert result.metrics["white_balance_reliability"] == "unavailable"
    assert (
        result.metrics["white_balance_note"]
        == "postprocessed_srgb_cutout_no_neutral_reference"
    )


def test_source_metadata_is_reported_without_changing_raw_rgb():
    colour = (224, 207, 188)
    exif = Image.Exif()
    exif[274] = 1
    result = extract_color_metrics(
        png_bytes(transparent_flower(colour), exif=exif), morph_kernel=0
    )

    assert result.metrics["source_exif_present"] is True
    assert result.metrics["camera_exif_present"] is False
    assert result.metrics["white_balance_metadata_present"] is False
    assert result.metrics["source_srgb_chunk_present"] is True
    assert result.metrics["source_icc_profile_present"] is False
    assert result.metrics["source_colour_space"] == "sRGB"
    assert tuple(result.metrics[channel] for channel in "RGB") == colour
    assert result.metrics["white_balance_applied"] is False
    assert result.metrics["white_balance_method"] == "none"


def test_srgb_to_cielab_matches_d65_reference_colours():
    lab = srgb_to_cielab(
        np.asarray([[255, 255, 255], [0, 0, 0], [255, 0, 0]], dtype=float)
    )
    assert lab[0] == pytest.approx([100.0, 0.0, 0.0], abs=2e-4)
    assert lab[1] == pytest.approx([0.0, 0.0, 0.0], abs=1e-12)
    assert lab[2] == pytest.approx([53.2408, 80.0925, 67.2032], abs=2e-3)


def test_joint_lab_peak_rejects_black_and_white_outliers():
    target = np.asarray([196, 72, 158], dtype=np.uint8)
    pixels = np.vstack(
        [
            np.repeat(target[None, :], 800, axis=0),
            np.zeros((100, 3), dtype=np.uint8),
            np.full((100, 3), 255, dtype=np.uint8),
        ]
    )
    result = estimate_joint_lab_peak(pixels)

    assert [result[f"peak_{channel}"] for channel in "RGB"] == target.tolist()
    assert result["peak_fraction"] == pytest.approx(0.8)
    assert result["second_peak_ratio"] < 0.5
    assert result["multimodal_colour"] is False


def test_joint_lab_peak_is_deterministic_for_noisy_target_and_rgb_lab_agree():
    random = np.random.default_rng(42)
    target = np.asarray([184, 92, 152], dtype=float)
    pixels = np.clip(random.normal(target, 6.0, size=(2000, 3)), 0, 255).round()
    first = estimate_joint_lab_peak(pixels)
    second = estimate_joint_lab_peak(pixels[random.permutation(len(pixels))])

    for key in (
        "peak_R",
        "peak_G",
        "peak_B",
        "peak_L",
        "peak_a",
        "peak_b",
        "peak_fraction",
        "second_peak_ratio",
    ):
        assert first[key] == pytest.approx(second[key], abs=1e-12)
    peak_rgb = np.asarray([[first[f"peak_{channel}"] for channel in "RGB"]])
    converted = srgb_to_cielab(peak_rgb)[0]
    assert converted == pytest.approx(
        [first["peak_L"], first["peak_a"], first["peak_b"]], abs=1e-10
    )
    assert np.linalg.norm(peak_rgb[0] - target) < 6
    assert all(0 <= first[f"peak_{channel}"] <= 255 for channel in "RGB")


def test_joint_lab_peak_flags_separated_bimodal_colour():
    first = np.repeat(np.asarray([[215, 55, 165]], dtype=np.uint8), 500, axis=0)
    second = np.repeat(np.asarray([[80, 165, 225]], dtype=np.uint8), 500, axis=0)
    result = estimate_joint_lab_peak(np.vstack([first, second]))

    assert result["second_peak_ratio"] == pytest.approx(1.0)
    assert result["multimodal_colour"] is True


def test_dual_alpha_and_hsv_peaks_expose_warm_exclusion_without_correction():
    rgba = np.empty((20, 20, 4), dtype=np.uint8)
    rgba[:, :, :] = (230, 180, 80, 255)  # Warm/yellow: excluded by legacy HSV mask.
    rgba[14:, :, :] = (210, 70, 170, 255)
    result = extract_color_metrics(png_bytes(rgba), morph_kernel=0)

    assert result.metrics["excluded_warm_fraction_visible"] == pytest.approx(0.7)
    assert result.metrics["hsv_peak_R"] == 210
    assert result.metrics["alpha_peak_R"] == 230
    assert result.metrics["alpha_hsv_peak_deltaE"] > 10
    assert "possible_warm_cast_or_petal_exclusion" in result.metrics["qc_flags"]
    assert "alpha_hsv_peak_disagreement" in result.metrics["qc_flags"]
    assert "low_mask_coverage" in result.metrics["qc_flags"]
    assert result.metrics["low_mask_coverage_threshold"] == 0.8
    assert result.metrics["white_balance_applied"] is False


def test_mask_components_and_true_clipping_are_diagnostic_flags():
    mask = np.zeros((30, 30), dtype=bool)
    mask[2:12, 2:12] = True
    mask[17:27, 17:27] = True
    summary = mask_component_summary(mask)
    assert summary["mask_component_count"] == 2
    assert summary["mask_large_component_count"] == 2
    assert summary["mask_largest_component_fraction"] == pytest.approx(0.5)
    assert summary["possible_multiple_flowers"] is True

    rgba = transparent_flower((210, 70, 170))
    rgba[10:15, 10:30, :] = (255, 255, 255, 255)
    result = extract_color_metrics(png_bytes(rgba), morph_kernel=0)
    assert result.metrics["all_channels_clipped_fraction"] > 0.01
    assert result.metrics["possible_overexposure"] is True
    assert "possible_overexposure" in result.metrics["qc_flags"]
    assert result.metrics["near_white_fraction"] > 0.1
    assert "high_near_white_fraction" in result.metrics["qc_flags"]


def test_exposure_filtered_peak_is_a_documented_sensitivity_ablation():
    rgba = np.empty((30, 30, 4), dtype=np.uint8)
    rgba[:, :, :] = (255, 255, 255, 255)
    rgba[18:, :, :] = (205, 65, 165, 255)
    result = extract_color_metrics(png_bytes(rgba), morph_kernel=0)

    assert tuple(result.metrics[f"hsv_peak_{channel}"] for channel in "RGB") == (
        255,
        255,
        255,
    )
    assert tuple(
        result.metrics[f"hsv_exposure_filtered_peak_{channel}"]
        for channel in "RGB"
    ) == (205, 65, 165)
    assert result.metrics["exposure_filtered_pixels"] == 360
    assert result.metrics["exposure_filtered_fraction"] == pytest.approx(0.4)
    assert result.metrics["exposure_filtered_peak_available"] is True
    assert tuple(result.metrics[channel] for channel in "RGB") != (205, 65, 165)


def test_white_alpha_foreground_is_supported():
    result = extract_color_metrics(
        png_bytes(transparent_flower((230, 230, 230))), morph_kernel=0
    )
    assert result.metrics["mask_pixels"] == 400
    assert tuple(result.metrics[channel] for channel in "RGB") == (230, 230, 230)


def test_blue_violet_alpha_foreground_is_supported():
    result = extract_color_metrics(
        png_bytes(transparent_flower((100, 170, 230))), morph_kernel=0
    )
    assert result.metrics["mask_pixels"] == 400
    assert tuple(result.metrics[channel] for channel in "RGB") == (100, 170, 230)


def test_hue_wrap_and_mean_median_are_distinct_and_correct():
    rgba = np.empty((10, 10, 4), dtype=np.uint8)
    rgba[:, :, :] = (160, 0, 100, 255)
    rgba[:5, :5, :] = (240, 0, 180, 255)
    result = extract_color_metrics(png_bytes(rgba), morph_kernel=0)

    assert result.metrics["mask_pixels"] == 100
    assert result.metrics["mean_R"] == pytest.approx(180.0)
    assert result.metrics["mean_B"] == pytest.approx(120.0)
    assert result.metrics["median_R"] == 160.0
    assert result.metrics["median_B"] == 100.0

    red = np.zeros((8, 8, 4), dtype=np.uint8)
    red[:, :, :] = (220, 20, 20, 255)
    wrapped = extract_color_metrics(png_bytes(red), morph_kernel=0)
    assert wrapped.metrics["mask_pixels"] == 64


def test_no_mask_is_a_recorded_result_not_an_exception():
    green = np.zeros((24, 24, 3), dtype=np.uint8)
    green[:, :, 1] = 180
    result = extract_color_metrics(png_bytes(green, mode="RGB"), morph_kernel=0)

    assert result.metrics["qc_status"] == "no_mask"
    assert result.metrics["mask_pixels"] == 0
    assert all(math.isnan(result.metrics[channel]) for channel in "RGB")


def test_directory_order_and_duplicate_hash_flags_are_deterministic(tmp_path):
    images = tmp_path / "images"
    images.mkdir()
    content = png_bytes(transparent_flower())
    (images / "b.png").write_bytes(content)
    (images / "a.png").write_bytes(content)

    records = collect_directory_records(images)
    assert [record.observation_id for record in records] == ["a", "b"]

    rows = process_records(records, qc_dir=tmp_path / "qc")
    assert [row["observation_id"] for row in rows] == ["a", "b"]
    assert all(row["duplicate_image_sha256"] for row in rows)
    assert all("duplicate_image" in row["qc_flags"] for row in rows)
    assert all(Path(row["mask_path"]).is_file() for row in rows)
    assert all(Path(row["overlay_path"]).is_file() for row in rows)
    assert all(Path(row["qc_panel_path"]).is_file() for row in rows)


def test_csv_output_refuses_implicit_overwrite(tmp_path):
    output = tmp_path / "colour.csv"
    row = {"observation_id": "obs-1", "qc_status": "no_mask", "R": math.nan}
    write_results([row], output)
    with output.open(encoding="utf-8", newline="") as stream:
        written = list(csv.DictReader(stream))
    assert written[0]["observation_id"] == "obs-1"
    assert written[0]["R"] == ""
    with pytest.raises(FileExistsError):
        write_results([row], output)


def test_cli_writes_new_csv_and_qc_images(tmp_path):
    images = tmp_path / "images"
    images.mkdir()
    (images / "flower.png").write_bytes(png_bytes(transparent_flower()))
    output = tmp_path / "derived" / "colour.csv"
    qc = tmp_path / "derived" / "qc"

    assert main(
        [
            "--input-dir",
            str(images),
            "--output",
            str(output),
            "--qc-dir",
            str(qc),
        ]
    ) == 0
    assert output.is_file()
    assert len(list((qc / "masks").glob("*.png"))) == 1
    assert len(list((qc / "overlays").glob("*.png"))) == 1
    panels = list((qc / "panels").glob("*.png"))
    assert len(panels) == 1
    with Image.open(panels[0]) as panel:
        assert panel.width > panel.height


def test_cli_csv_and_qc_are_reproducible_when_timestamp_is_pinned(
    tmp_path, monkeypatch
):
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "1720000000")
    images = tmp_path / "images"
    images.mkdir()
    (images / "flower.png").write_bytes(png_bytes(transparent_flower()))
    output = tmp_path / "colour.csv"
    qc = tmp_path / "qc"
    arguments = [
        "--input-dir",
        str(images),
        "--output",
        str(output),
        "--qc-dir",
        str(qc),
    ]

    assert main(arguments) == 0
    first_csv = output.read_bytes()
    first_qc = {
        path.relative_to(qc): path.read_bytes()
        for path in sorted(qc.rglob("*.png"))
    }

    assert main([*arguments, "--overwrite"]) == 0
    assert output.read_bytes() == first_csv
    assert {
        path.relative_to(qc): path.read_bytes()
        for path in sorted(qc.rglob("*.png"))
    } == first_qc


def test_qc_sample_size_is_deterministic_and_does_not_render_every_image(tmp_path):
    images = tmp_path / "many-images"
    images.mkdir()
    for index in range(25):
        colour = (200 + index, 60, 160)
        (images / f"flower-{index:02d}.png").write_bytes(
            png_bytes(transparent_flower(colour))
        )
    records = collect_directory_records(images)
    first = process_records(
        records,
        qc_dir=tmp_path / "qc-first",
        qc_sample_size=20,
    )
    second = process_records(
        list(reversed(records)),
        qc_dir=tmp_path / "qc-second",
        qc_sample_size=20,
    )

    first_ids = {row["observation_id"] for row in first if row["qc_sampled"]}
    second_ids = {row["observation_id"] for row in second if row["qc_sampled"]}
    assert len(first_ids) == 20
    assert first_ids == second_ids
    assert len(list((tmp_path / "qc-first" / "masks").glob("*.png"))) == 20
    assert len(list((tmp_path / "qc-first" / "panels").glob("*.png"))) == 20


def test_qc_always_includes_no_mask_duplicates_and_flagged_rows(tmp_path):
    images = tmp_path / "stratified-images"
    images.mkdir()
    for index in range(30):
        colour = (180 + index, 45 + index % 10, 140 + index % 20)
        (images / f"flower-{index:02d}.png").write_bytes(
            png_bytes(transparent_flower(colour))
        )
    duplicate = png_bytes(transparent_flower((205, 55, 165)))
    (images / "duplicate-a.png").write_bytes(duplicate)
    (images / "duplicate-b.png").write_bytes(duplicate)
    green = np.zeros((40, 40, 3), dtype=np.uint8)
    green[:, :, 1] = 180
    (images / "no-mask.png").write_bytes(png_bytes(green, mode="RGB"))

    rows = process_records(
        collect_directory_records(images),
        qc_dir=tmp_path / "qc-stratified",
        qc_sample_size=20,
    )
    by_id = {row["observation_id"]: row for row in rows}
    assert sum(row["qc_sampled"] for row in rows) >= 20
    assert by_id["no-mask"]["qc_sampled"]
    assert by_id["no-mask"]["qc_status"] == "no_mask"
    assert "No pixels" in by_id["no-mask"]["qc_note"]
    for observation_id in ("duplicate-a", "duplicate-b"):
        assert by_id[observation_id]["qc_sampled"]
        assert by_id[observation_id]["qc_status"] == "manual_review_required"
        assert Path(by_id[observation_id]["qc_panel_path"]).is_file()
        assert "possible multiple flowers" in by_id[observation_id]["qc_note"]


def test_run_metadata_uses_one_source_date_epoch_timestamp(tmp_path, monkeypatch):
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "0")
    images = tmp_path / "metadata-images"
    images.mkdir()
    (images / "one.png").write_bytes(png_bytes(transparent_flower()))
    (images / "two.png").write_bytes(png_bytes(transparent_flower((220, 80, 180))))

    rows = process_records(
        collect_directory_records(images),
        qc_dir=tmp_path / "qc-metadata",
        qc_sample_size=None,
    )
    assert {row["processed_at"] for row in rows} == {"1970-01-01T00:00:00Z"}
    assert {row["extraction_version"] for row in rows} == {"2.2.2"}
    assert all(row["source_path"].endswith(".png") for row in rows)
    assert all(not Path(row["source_path"]).is_absolute() for row in rows)
    assert all("mean_R" in row and "median_R" in row for row in rows)
    assert all("R_mean" not in row and "R_median" not in row for row in rows)


def test_legacy_missing_and_top_twenty_differences_are_mandatory_qc(tmp_path):
    images = tmp_path / "legacy-images"
    images.mkdir()
    for index in range(60):
        colour = (170 + index, 45 + index % 5, 130 + index % 17)
        (images / f"legacy-{index:02d}.png").write_bytes(
            png_bytes(transparent_flower(colour))
        )
    base_records = collect_directory_records(images)
    records = []
    for index, record in enumerate(base_records):
        current_r = 170 + index
        current_g = 45 + index % 5
        current_b = 130 + index % 17
        if index >= 55:
            records.append(record)
        else:
            records.append(
                replace(
                    record,
                    legacy_R=current_r + index,
                    legacy_G=current_g,
                    legacy_B=current_b,
                )
            )

    rows = process_records(
        records,
        qc_dir=tmp_path / "qc-legacy",
        qc_sample_size=20,
        processed_at="2000-01-01T00:00:00Z",
    )
    # 25 mandatory rows (5 missing + top-20 deltas) plus 20 stratified rows.
    assert sum(row["qc_sampled"] for row in rows) == 45
    by_id = {row["observation_id"]: row for row in rows}
    for index in range(35, 55):
        row = by_id[f"legacy-{index:02d}"]
        assert row["qc_sampled"]
        assert row["delta_R_vs_legacy"] == pytest.approx(-index)
        assert row["delta_G_vs_legacy"] == pytest.approx(0)
        assert row["delta_B_vs_legacy"] == pytest.approx(0)
        assert row["delta_rgb_euclidean_vs_legacy"] == pytest.approx(index)
        assert "top20_legacy_rgb_difference" in row["qc_note"]
    for index in range(55, 60):
        row = by_id[f"legacy-{index:02d}"]
        assert row["qc_sampled"]
        assert math.isnan(row["delta_rgb_euclidean_vs_legacy"])
        assert "legacy_rgb_all_missing" in row["qc_note"]


def test_cli_rejects_qc_samples_smaller_than_twenty(tmp_path):
    images = tmp_path / "images"
    images.mkdir()
    (images / "flower.png").write_bytes(png_bytes(transparent_flower()))
    with pytest.raises(SystemExit, match="at least 20"):
        main(
            [
                "--input-dir",
                str(images),
                "--output",
                str(tmp_path / "colour.csv"),
                "--qc-dir",
                str(tmp_path / "qc"),
                "--qc-sample-size",
                "19",
            ]
        )


def test_standard_drawing_workbook_maps_image_to_observation(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.drawing.image import Image as ExcelImage

    image_path = tmp_path / "flower.png"
    image_path.write_bytes(png_bytes(transparent_flower()))
    workbook_path = tmp_path / "input.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "observations"
    sheet.append(
        [
            "observation_id",
            "url",
            "date",
            "latitude",
            "longitude",
            "petal",
            "R",
            "G",
            "B",
        ]
    )
    sheet.append(
        ["obs-001", "https://example.test/1", "2025-07-01", 35.0, 138.0, None, 1, 2, 3]
    )
    sheet.add_image(ExcelImage(str(image_path)), "F2")
    workbook.save(workbook_path)

    records = collect_workbook_records(
        workbook_path, sheet_name="observations", image_column="petal"
    )
    assert len(records) == 1
    record = records[0]
    assert record.observation_id == "obs-001"
    assert record.source_row == 2
    assert record.url == "https://example.test/1"
    assert record.latitude == 35.0
    assert record.longitude == 138.0
    assert record.legacy_R == 1
    assert record.zip_member and record.zip_member.startswith("xl/media/")

    rows = process_records(records, qc_dir=tmp_path / "qc")
    assert rows[0]["qc_status"] == "ok"
    assert rows[0]["image_sha256"]


def test_xlsx_output_is_optional_and_readable(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    output = tmp_path / "colour.xlsx"
    write_results([{"observation_id": "obs-1", "R": 123.5}], output)
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook["color_v2"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[0][0] == "observation_id"
    assert rows[1][0] == "obs-1"


def test_rich_value_images_follow_vm_metadata_and_can_be_reused(tmp_path):
    workbook_path = tmp_path / "rich.xlsx"
    image = png_bytes(transparent_flower())
    parts = {
        "xl/workbook.xml": """
          <workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                    xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
            <sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>
          </workbook>""",
        "xl/_rels/workbook.xml.rels": """
          <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
            <Relationship Id="rId1" Type="worksheet" Target="worksheets/sheet1.xml"/>
          </Relationships>""",
        "xl/worksheets/sheet1.xml": """
          <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
            <sheetData>
              <row r="2"><c r="E2" vm="1" t="e"><v>#VALUE!</v></c></row>
              <row r="3"><c r="E3" vm="2" t="e"><v>#VALUE!</v></c></row>
            </sheetData>
          </worksheet>""",
        "xl/metadata.xml": """
          <metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                    xmlns:xlrd="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata">
            <futureMetadata name="XLRICHVALUE" count="2">
              <bk><extLst><ext><xlrd:rvb i="0"/></ext></extLst></bk>
              <bk><extLst><ext><xlrd:rvb i="1"/></ext></extLst></bk>
            </futureMetadata>
            <valueMetadata count="2">
              <bk><rc t="1" v="0"/></bk><bk><rc t="1" v="1"/></bk>
            </valueMetadata>
          </metadata>""",
        "xl/richData/rdrichvaluestructure.xml": """
          <rvStructures xmlns="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata">
            <s t="_localImage"><k n="_rvRel:LocalImageIdentifier" t="i"/><k n="CalcOrigin" t="i"/></s>
          </rvStructures>""",
        "xl/richData/rdrichvalue.xml": """
          <rvData xmlns="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata">
            <rv s="0"><v>0</v><v>5</v></rv><rv s="0"><v>0</v><v>5</v></rv>
          </rvData>""",
        "xl/richData/richValueRel.xml": """
          <richValueRels xmlns="http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
                         xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
            <rel r:id="rId1"/>
          </richValueRels>""",
        "xl/richData/_rels/richValueRel.xml.rels": """
          <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
            <Relationship Id="rId1" Type="image" Target="../media/image1.png"/>
          </Relationships>""",
    }
    with zipfile.ZipFile(workbook_path, "w") as archive:
        for member, text in parts.items():
            archive.writestr(member, text.strip())
        archive.writestr("xl/media/image1.png", image)

    sheet_name, mapping = workbook_image_map(workbook_path)
    assert sheet_name == "Sheet1"
    assert mapping == {"E2": "xl/media/image1.png", "E3": "xl/media/image1.png"}


def test_duplicate_observation_ids_in_workbook_are_rejected(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    workbook_path = tmp_path / "duplicate_ids.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["observation_id", "petal"])
    sheet.append(["same", None])
    sheet.append(["same", None])
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Duplicate observation_id"):
        collect_workbook_records(workbook_path)
